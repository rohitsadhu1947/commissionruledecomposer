"""
POC: Two-Wheeler pay-in rate extractor for the SBI General grid.

Demonstrates the pipeline for the EASIEST category (TW):
  - reads the 'PCV, MISD & TW' sheet
  - corrects the global -2 column offset (data sits 2 cols left of header labels)
  - emits canonical RATE rules (see ARCHITECTURE.md schema)
  - emits validation warnings for anomalies it cannot resolve confidently

Insurer-specific bits (sheet name, offset, TW column map, grid_version) are
declared in CONFIG up top -- in the real system this is the per-insurer
mapping config, not code.
"""
import json
import re
import openpyxl
from openpyxl.utils import get_column_letter

FILE = "Provincial sbi Grid revision__May'26 _w.e.f 11th May'26.xlsx"

CONFIG = {
    "insurer": "SBI_GENERAL",
    "grid_version": {
        "effective_from": "2026-05-11",
        "effective_to": "2026-05-31",
        "source_file": FILE,
        "source_sheet": "PCV, MISD & TW",
    },
    "header_row_for_geo": 4,   # Region | State Name | Circle
    "first_data_row": 5,
    "last_data_row": 49,
    # On the TW sheet the matching geography is STATE (col 2). 'Circle' (col 3) is just a
    # grouping label kept as metadata, NOT a match dimension. (Other sheets key on rto_cluster.)
    "geo_cols": {"region": 1, "state": 2, "circle": 3},
    # Headers sit at their own columns (NO offset). Verified against the file:
    #   Z/AA = School Bus, AB/AC = Pvt Car, AD/AE/AF = 2 Wheeler.
    "tw_columns": [
        {"col": 30, "sub_segment": "SCOOTER", "cc_band": {"min_cc": 0, "max_cc": 150},
         "policy_type": ["COMP", "SAOD"], "raw_header": "Scooter upto 150 cc (Comp & SAOD)"},
        {"col": 31, "sub_segment": "BIKE", "cc_band": {"min_cc": 0, "max_cc": 125},
         "policy_type": ["COMP", "SAOD"], "raw_header": "Bike upto 125 cc"},
        {"col": 32, "sub_segment": "BIKE", "cc_band": {"min_cc": 125, "max_cc": None},
         "policy_type": ["COMP", "SAOD"], "raw_header": "C. Above 125cc"},
    ],
    "term": "1+1",
}


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return float(s)
    return None


def extract():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb[CONFIG["source_sheet"]] if "source_sheet" in CONFIG else wb[CONFIG["grid_version"]["source_sheet"]]
    rules, warnings = [], []
    rid = 0
    for r in range(CONFIG["first_data_row"], CONFIG["last_data_row"] + 1):
        region = ws.cell(row=r, column=CONFIG["geo_cols"]["region"]).value
        state = ws.cell(row=r, column=CONFIG["geo_cols"]["state"]).value
        circle = ws.cell(row=r, column=CONFIG["geo_cols"]["circle"]).value
        if not (state and str(state).strip()):
            continue
        circle = str(circle).strip() if circle else None
        state = str(state).strip()

        for spec in CONFIG["tw_columns"]:
            val = num(ws.cell(row=r, column=spec["col"]).value)
            if val is None:
                continue  # blank = not offered in this cluster (NOT a zero rate)
            rid += 1
            rules.append({
                "rule_id": f"SBI-TW-{rid:04d}",
                "insurer": CONFIG["insurer"],
                "grid_version": CONFIG["grid_version"],
                "rule_type": "RATE",
                "scope": {
                    "category": "TW",
                    "sub_segment": spec["sub_segment"],
                    "cc_band": spec["cc_band"],
                    "policy_type": spec["policy_type"],
                    "rto_cluster": None,          # TW keys on state; cluster N/A for this sheet
                    "state": state,
                    "circle": circle,             # metadata only, not a match dimension
                    "region": str(region).strip() if region else None,
                    "term": CONFIG["term"],
                },
                "effect": {"value": val, "applies_on": "NET"},
                "precedence": 100,
                "source": {
                    "cell": f"{get_column_letter(spec['col'])}{r}",
                    "raw_header": spec["raw_header"],
                    "raw_row_label": state,
                    "confidence": 1.0,
                    "review_status": "PENDING",
                },
            })
        # validation: scooter rate present but bike (<=125) blank in same state (e.g. Karnataka 50/blank)
        scooter = num(ws.cell(row=r, column=30).value)
        bike = num(ws.cell(row=r, column=31).value)
        if scooter is not None and bike is None:
            warnings.append({
                "row": r, "state": state, "cell": f"AD{r}",
                "issue": f"Scooter <=150 has a rate ({scooter}) but Bike <=125 (AE) is blank -- confirm bike is intentionally not offered",
            })

    # sheet-level validation: 'Above 125cc' (AF) is empty for every row
    af_vals = [num(ws.cell(row=r, column=32).value)
               for r in range(CONFIG["first_data_row"], CONFIG["last_data_row"] + 1)]
    if not any(v is not None for v in af_vals):
        warnings.append({
            "scope": "sheet", "cell": "AF column",
            "issue": "'C. Above 125cc' column is empty for all rows -- no rate for bikes >125cc on this sheet; "
                     "confirm whether >125cc TW is declined or specified elsewhere",
        })
    return rules, warnings


if __name__ == "__main__":
    rules, warnings = extract()
    out = {"rules": rules, "validation_warnings": warnings,
           "summary": {"rate_rules": len(rules), "warnings": len(warnings)}}
    with open("poc_tw_rules.json", "w") as f:
        json.dump(out, f, indent=2)
    print(f"Extracted {len(rules)} TW RATE rules, {len(warnings)} warnings -> poc_tw_rules.json")
    for s in rules[:4]:
        print(f"  {s['rule_id']}: {s['scope']['sub_segment']} {s['scope']['cc_band']} "
              f"@ {s['scope']['state']}/{s['scope']['rto_cluster']} = {s['effect']['value']}% (cell {s['source']['cell']})")
    print("  ...")
    for w in warnings:
        print(f"  WARN {w.get('cell')} {w.get('state', w.get('scope',''))}: {w['issue']}")

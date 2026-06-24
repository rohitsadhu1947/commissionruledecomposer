"""
Build the COMPLETE, ATOMIC rule catalog the team loads into the commission platform.

Output (one row == one platform rule, fully expanded; no list-valued conditions):
  rules_catalog.xlsx   - formatted spreadsheet for human review / sign-off
  rules_catalog.csv    - same, portable
  rules_catalog.json   - same rows as objects, for API/bulk import

Design choices (per requirements):
  - geography kept AS-IS from the source (grid uses city/region labels like
    "Mumbai", "RO Maharashtra", "NCR"; notes use state names). We add a
    `canonical_state` helper column (folded via the RTO master) for cross-ref,
    but `geo_label` is the authoritative source value.
  - atomic expansion: every list-valued condition (policy_type, sub_segment,
    make, model, states, allowed_states, allowed_rto_codes) is exploded so each
    row carries exactly one value. ALLOW-ONLY notes become explicit ALLOW rows
    for each listed geo + one DECLINE_DEFAULT row for the segment.
"""
import csv
import json

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from poc_tw_extractor import extract
from geo_normalize import to_state

from catalog_schema import COLUMNS, row as _schema_row

INSURER = "SBI_GENERAL"


def _geo_kind(label, is_rto=False):
    if is_rto:
        return "RTO_CODE"
    if label is None:
        return ""
    up = label.upper()
    if up == "PAN_INDIA":
        return "PAN_INDIA"
    # crude: real city if RTO master maps the literal as a city, region if "RO"/"REST",
    # otherwise treat as state. canonical_state handles the actual fold.
    if up.startswith("RO ") or up.startswith("ROK") or up.startswith("REST OF") or "(" in label:
        return "REGION"
    return "STATE_OR_CITY"


def row(**kw):
    kw.setdefault("insurer", INSURER)
    return _schema_row(**kw)


# ---------------------------------------------------------------- RATE rules
def expand_rate(rate_rules):
    out = []
    for rr in rate_rules:
        sc, eff, src = rr["scope"], rr["effect"], rr["source"]
        cc = sc.get("cc_band") or {}
        label = sc.get("state")
        for pol in sc.get("policy_type") or [None]:
            out.append(row(
                catalog_id=f'{rr["rule_id"]}-{pol or "ANY"}',
                source_rule_id=rr["rule_id"],
                rule_type="RATE", effect="RATE",
                pay_in_pct=eff["value"], applies_on=eff.get("applies_on", "NET"),
                category=sc.get("category"), sub_segment=sc.get("sub_segment"),
                cc_min=cc.get("min_cc"), cc_max=cc.get("max_cc"),
                policy_type=pol,
                geo_kind=_geo_kind(label), geo_label=label,
                canonical_state=to_state(label) or "",
                source_sheet=rr.get("grid_version", {}).get("source_sheet", ""),
                source_cell=src.get("cell"), source_text=src.get("raw_header", ""),
                confidence=src.get("confidence"), review_status=src.get("review_status"),
            ))
    return out


# --------------------------------------------------------- ELIGIBILITY rules
def _src(r):
    s = r.get("source", {})
    return dict(source_sheet=s.get("sheet", ""), source_text=s.get("source_text", ""),
                confidence=s.get("confidence"), review_status=s.get("review_status", "PENDING"))


def _as_list(v):
    return v if isinstance(v, list) else [v]


def expand_eligibility(elig_rules):
    out = []
    n = 0
    for r in elig_rules:
        rid, sc, eff = r["rule_id"], r["scope"], r["effect"]
        base = dict(source_rule_id=rid, rule_type="ELIGIBILITY",
                    category=sc.get("category"), **_src(r))
        subs = _as_list(sc.get("sub_segment")) if sc.get("sub_segment") else [""]
        pols = _as_list(sc.get("policy_type")) if sc.get("policy_type") else [""]

        def emit(**kw):
            nonlocal n
            n += 1
            out.append(row(catalog_id=f"{rid}-{n:03d}", **{**base, **kw}))

        # --- ALLOW_ONLY: explicit ALLOW rows per listed geo + one DECLINE_DEFAULT
        if eff.get("mode") == "ALLOW_ONLY":
            for sub in subs:
                for pol in pols:
                    for st in eff.get("allowed_states", []):
                        emit(effect="ALLOW", sub_segment=sub, policy_type=pol,
                             geo_kind=_geo_kind(st), geo_label=st,
                             canonical_state=to_state(st) or "",
                             reason="Doable geo (allow-list)")
                    for code in eff.get("allowed_rto_codes", []):
                        emit(effect="ALLOW", sub_segment=sub, policy_type=pol,
                             geo_kind="RTO_CODE", rto_code=code, geo_label=code,
                             canonical_state=to_state(code) or "",
                             reason="Doable RTO (allow-list)")
                    emit(effect="DECLINE_DEFAULT", sub_segment=sub, policy_type=pol,
                         geo_kind="ELSEWHERE", geo_label="* (not in allow-list)",
                         reason=eff.get("reason", "Not in allow-list"))
            continue

        reason = eff.get("reason", "")

        # --- make + model decline (optionally state-gated)  e.g. TW2, TW1
        if "make_models" in sc:
            states = sc.get("states", [None])
            for sub in subs:
                for pol in pols:
                    for st in states:
                        for mm in sc["make_models"]:
                            for model in (mm.get("models") or ["ALL"]):
                                emit(effect="DECLINE", sub_segment=sub or "BIKE",
                                     policy_type=pol, make=mm["make"], model=model,
                                     match=mm.get("match", ""),
                                     geo_kind=_geo_kind(st) if st else "PAN_INDIA",
                                     geo_label=st if st else "PAN_INDIA",
                                     canonical_state=(to_state(st) or "") if st else "",
                                     reason=reason)
            continue

        # --- make x state declines  e.g. TW5
        if "make_state_declines" in sc:
            for sub in subs:
                for pol in pols:
                    for msd in sc["make_state_declines"]:
                        for st in msd["states"]:
                            emit(effect="DECLINE", sub_segment=sub, policy_type=pol,
                                 make=msd["make"],
                                 model=("ALL OTHER MAKES" if msd["make"] == "*OTHER*" else "ALL"),
                                 geo_kind=_geo_kind(st), geo_label=st,
                                 canonical_state=to_state(st) or "",
                                 reason=reason)
            continue

        # --- plain state-list decline  e.g. TW6
        if "states" in sc:
            for sub in subs:
                for pol in pols:
                    for st in sc["states"]:
                        emit(effect="DECLINE", sub_segment=sub, policy_type=pol,
                             geo_kind=_geo_kind(st), geo_label=st,
                             canonical_state=to_state(st) or "",
                             reason=reason)
            continue

        # --- age gate (no geo/make)  e.g. G2-COMP, G2-SATP
        # The note semantics are "allowed up to N years" -> decline only ABOVE N.
        # _in_band is inclusive on the low edge, so the decline band starts at
        # N+1 (F10: avoids declining a vehicle of exactly N years, which is allowed).
        if "age_years_min" in sc:
            for pol in pols:
                emit(effect="DECLINE", sub_segment=(subs[0] if subs[0] else ""),
                     policy_type=pol, age_min=sc["age_years_min"] + 1,
                     geo_kind="PAN_INDIA", geo_label="PAN_INDIA", reason=reason)
            continue

    return out


# -------------------------------------------------------------------- writers
def write_csv(rows, path):
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=COLUMNS)
        w.writeheader()
        w.writerows(rows)


def write_json(rows, meta, path):
    with open(path, "w") as f:
        json.dump({"meta": meta, "rules": rows}, f, indent=2)


def write_xlsx(rows, meta, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rule Catalog"
    head_fill = PatternFill("solid", fgColor="1F2937")
    head_font = Font(bold=True, color="FFFFFF")
    rate_fill = PatternFill("solid", fgColor="E8F5E9")
    decline_fill = PatternFill("solid", fgColor="FDE8E8")
    allow_fill = PatternFill("solid", fgColor="FFF8E1")

    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(1, c)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r[c] for c in COLUMNS])
        last = ws.max_row
        eff = r["effect"]
        fill = rate_fill if r["rule_type"] == "RATE" else (
            allow_fill if eff == "ALLOW" else decline_fill)
        for c in range(1, len(COLUMNS) + 1):
            ws.cell(last, c).fill = fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{ws.max_row}"
    widths = {"reason": 42, "source_text": 48, "geo_label": 22, "canonical_state": 18,
              "model": 18, "make": 16, "catalog_id": 20, "source_rule_id": 18}
    for i, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)

    # cover/meta sheet
    cov = wb.create_sheet("README", 0)
    cov["A1"] = "SBI General — Two-Wheeler Commission Rule Catalog"
    cov["A1"].font = Font(bold=True, size=14)
    lines = [
        "", f"Insurer: {meta['insurer']}",
        f"Grid version: {meta['grid_version']['effective_from']} to {meta['grid_version']['effective_to']}",
        f"Source file: {meta['source_file']}",
        f"Generated rows: {meta['counts']['total']} "
        f"(RATE {meta['counts']['rate']}, ELIGIBILITY {meta['counts']['eligibility']})",
        "",
        "Each row = one atomic platform rule. Geography is the SOURCE label",
        "(geo_label); canonical_state is the RTO-master fold for cross-reference.",
        "effect: RATE (pay_in_pct applies) | DECLINE | ALLOW | DECLINE_DEFAULT.",
        "review_status = PENDING means a human must confirm before go-live.",
    ]
    for i, t in enumerate(lines, 2):
        cov[f"A{i}"] = t
    cov.column_dimensions["A"].width = 90

    wb.save(path)


def main():
    rate_rules, warnings = extract()
    with open("poc_tw_notes_rules.json") as f:
        notes = json.load(f)
    elig_rules = notes["eligibility_rules"]

    rate_rows = expand_rate(rate_rules)
    elig_rows = expand_eligibility(elig_rules)
    rows = rate_rows + elig_rows

    meta = {
        "insurer": INSURER,
        "category": "TW",
        "grid_version": notes.get("grid_version", {}),
        "source_file": "Provincial sbi Grid revision__May'26 _w.e.f 11th May'26.xlsx",
        "counts": {"total": len(rows), "rate": len(rate_rows), "eligibility": len(elig_rows)},
        "geography": "source labels preserved; canonical_state via RTO master",
        "expansion": "atomic (one condition per row)",
        "warnings": warnings,
    }

    write_csv(rows, "rules_catalog.csv")
    write_json(rows, meta, "rules_catalog.json")
    write_xlsx(rows, meta, "rules_catalog.xlsx")

    print(f"Rule catalog generated: {len(rows)} atomic rules")
    print(f"  RATE        : {len(rate_rows)}")
    print(f"  ELIGIBILITY : {len(elig_rows)}")
    print("  -> rules_catalog.xlsx / .csv / .json")
    if warnings:
        print(f"  ({len(warnings)} extraction warnings carried in JSON meta)")


if __name__ == "__main__":
    main()

"""
Extractor for the Cholamandalam MS retail-broking motor grid
("June'26 Grid - Retail Broking Motor Cholamandalam.xlsx").

Emits rows in the same 27-column atomic-catalog schema as the SBI extractors
(one source cell == one catalog row), insurer = CHOLA_MS.

Source layout (very different from SBI):
  - values are FRACTIONS (0.425 == 42.5%) -> x100 on extraction
  - geography is multi-state group labels in the column headers
    ("AP/TS", "GJ/DN/DD", "AS/ML/TR/AR/NL/SK", zone labels "TN-Chennai", "ROM")
  - 'TW June 26'           state blocks x make-groups x cc bands; row types are
                           NB volume slabs (NOP), ANNUAL renewal, ACT (=SATP)
  - 'Private car June-26'  PROD/SUBCLASS rows x state-group cols x ACT/PACK
  - ' CV June 26'          subclass rows x state-group cols x ACT/PACK
  - 'conditions'           premium-basis notes + the "TW SAOD blocked" rule
  - 'RTO '                 MH + TN zone splits (Mumbai/Pune/Central MH, Chennai)

Policy vocabulary normalization: PACK->COMP, ACT->SATP, SOD->SAOD; the TW NB
volume slabs stay visible in policy_type ("COMP-NEW (100-500 NOP)") because
they price differently.
"""
import glob
import re

import openpyxl
from openpyxl.utils import get_column_letter

from catalog_schema import row
import geo_lookup as gn

INSURER = "CHOLA_MS"


def _file():
    return glob.glob("June'26 Grid - Retail Broking Motor Cholamandalam*.xlsx")[0]
SRC_TW = "TW June 26"
SRC_PC = "Private car June-26"
SRC_CV = " CV June 26"
SRC_RTO = "RTO "

# --- geo: state-group label -> canonical states ------------------------------
_TOKEN_ALIAS = {
    "AP": "ANDHRA PRADESH", "TS": "TELANGANA", "TN": "TAMIL NADU",
    "KA": "KARNATAKA", "KL": "KERALA", "LD": "LAKSHADWEEP", "PY": "PONDICHERRY",
    "MH": "MAHARASHTRA", "GJ": "GUJARAT", "DN": "DADRA & NAGAR HAVELI",
    "DD": "DAMAN & DIU", "GA": "GOA", "BR": "BIHAR", "BH": "BIHAR",
    "JH": "JHARKHAND", "CG": "CHATTISGARH", "OD": "ORISSA", "WB": "WEST BENGAL",
    "AS": "ASSAM", "ML": "MEGHALAYA", "TR": "TRIPURA", "AR": "ARUNACHAL PRADESH",
    "NL": "NAGALAND", "SK": "SIKKIM", "DL": "DELHI", "HR": "HARYANA",
    "HP": "HIMACHAL PRADESH", "JK": "JAMMU AND KASHMIR", "PB": "PUNJAB",
    "MP": "MADHYA PRADESH", "RJ": "RAJASTHAN", "UK": "UTTARAKHAND",
    "UP": "UTTAR PRADESH", "AN": "ANDAMAN & NICOBAR ISLANDS",
    "ROK": "KARNATAKA", "ROTN": "TAMIL NADU", "ROM": "MAHARASHTRA",
    "MUMBAI": "MAHARASHTRA", "PUNE": "MAHARASHTRA", "CENTRAL MH": "MAHARASHTRA",
    "CHENNAI": "TAMIL NADU", "EAST UP": "UTTAR PRADESH", "UP- EAST": "UTTAR PRADESH",
}


def label_states(label):
    """'GJ/DN/DD' -> {GUJARAT, DADRA..., DAMAN...}; 'KA\\n(Bangalore)' -> {KARNATAKA}."""
    if not label:
        return set()
    txt = re.sub(r"\([^)]*\)", "", str(label)).replace("\n", " ").strip()
    out = set()
    for part in re.split(r"[/]", txt):
        p = part.strip().upper().rstrip("-").strip()
        if not p:
            continue
        st = (_TOKEN_ALIAS.get(p) or _TOKEN_ALIAS.get(p.replace(" ", ""))
              or _TOKEN_ALIAS.get(p.split("-")[0].strip()) or gn.to_state(p))
        if st:
            out.add(st)
    return out


def _geo(label):
    sts = label_states(label)
    return {"geo_kind": "STATE_GROUP", "geo_label": re.sub(r"\s+", " ", str(label).strip()),
            "canonical_state": next(iter(sts)) if len(sts) == 1 else ""}


def pct(v):
    """Fractional cell -> percent number, else None (blank / text)."""
    if v is None or isinstance(v, str):
        return None
    v = float(v)
    return round(v * 100, 3) if -1.0 <= v <= 1.0 else round(v, 3)


# --- TW sheet specs ----------------------------------------------------------
# (first col, make label, [band labels...]); bands: (label, sub_segment, cc_min, cc_max)
_BANDS4 = [("150cc", "TW MC <=150CC", 0, 150), ("SCOOTER", "TW SCOOTER", None, None),
           ("150_350cc", "TW MC 150-350CC", 151, 350), ("350cc", "TW MC >350CC", 351, None)]
TW_MAKES = [
    (4,  "HERO",   _BANDS4), (9, "TVS", _BANDS4), (14, "SUZUKI", _BANDS4),
    (19, "YAMAHA", _BANDS4), (24, "ROYAL ENFIELD", _BANDS4[2:]),
    (27, "HONDA",  _BANDS4), (32, "BAJAJ", _BANDS4), (37, "OTHER MAKES", _BANDS4),
]
TW_EV = [(44, "TW EV BIKE <=7KW", None, None), (45, "TW EV SCOOTER <=7KW", None, None),
         (46, "TW EV 7-16KW", None, None), (47, "TW EV >16KW", None, None)]
# Per-row Type label (col C / EV col AQ) -> policy_type. Driven by the actual
# label, NOT a fixed offset, because some state blocks have an extra 'SOD' row
# (e.g. GA = 6 rows) which shifts a fixed-stride loop and drops every later state.
def _tw_policy(type_label):
    t = re.sub(r"\s+", " ", str(type_label or "").strip()).upper()
    if t.startswith("NEW"):
        return "COMP-" + str(type_label).strip()      # COMP-NEW (100-500 NOP) etc.
    if t == "ANNUAL":
        return "COMP-RENEWAL"
    if t == "SOD":
        return "SAOD"
    if t == "ACT":
        return "SATP"
    return None  # header/footer/blank row

# --- PC sheet specs: (row, policy, sub suffix, cc_min, cc_max) ---------------
_CC3 = [("UPTO 1000CC", 0, 1000), ("1000-1500CC", 1001, 1500), ("ABOVE 1500CC", 1501, None)]


def _pc_rows():
    out = []
    for i, (cc_lbl, lo, hi) in enumerate(_CC3):
        out.append((7 + i,  "COMP", f"PC {cc_lbl}", "", lo, hi))
        out.append((12 + i, "SAOD", f"PC {cc_lbl} [PETROL]", "", lo, hi))
        out.append((16 + i, "SAOD", f"PC {cc_lbl} [DIESEL]", "", lo, hi))
        out.append((21 + i, "COMP", f"PC {cc_lbl} [NCB>=25%]", "", lo, hi))
        out.append((26 + i, "SAOD", f"PC {cc_lbl} [PETROL] [NCB>=25%]", "", lo, hi))
        out.append((30 + i, "SAOD", f"PC {cc_lbl} [DIESEL] [NCB>=25%]", "", lo, hi))
    return out


# --- CV sheet specs: (row, category, sub_segment) ---------------------------
# make (col E) and model (col F) are READ from the sheet at extraction time, not
# hardcoded here, so make/model lists stay faithful and a monthly re-upload picks
# up any changes. Only the friendly sub_segment label is mapped per row.
CV_ROWS = [
    (6,  "GCV",  "GCV 3W"),
    (7,  "GCV",  "GCV 3W [ELECTRIC]"),
    (8,  "GCV",  "GCV UPTO 3.5T"),
    (9,  "GCV",  "GCV UPTO 3.5T"),
    (10, "GCV",  "GCV UPTO 3.5T [ELECTRIC]"),
    (11, "GCV",  "GCV 3.5-7.5T"),
    (12, "GCV",  "GCV 7.5-12T"),
    (13, "GCV",  "GCV 12-16T"),
    (14, "GCV",  "GCV 16-20T"),
    (15, "GCV",  "GCV 20-40T"),
    (16, "GCV",  "GCV 40-43T"),
    (17, "GCV",  "GCV 43-47.5T"),
    (18, "GCV",  "GCV 47.5-56T"),
    (20, "PCV",  "PCV <6 3W AUTO"),
    (21, "PCV",  "PCV <6 3W AUTO [ELECTRIC]"),
    (22, "PCV",  "PCV <6 4W <1500CC"),
    (23, "PCV",  "PCV <6 4W >1500CC"),
    (25, "PCV",  "PCV 6+ SCHOOL BUS"),
    (26, "PCV",  "PCV 6+ STAFF BUS"),
    (27, "PCV",  "PCV 6+ BIG TAXIS"),
    (28, "PCV",  "PCV 6+ MAXI CAB"),
    (29, "PCV",  "PCV 6+ BUS"),
    (31, "MISD", "MISD TRACTOR [NEW]"),
    (32, "MISD", "MISD TRACTOR [RENEWAL]"),
    (33, "MISD", "MISD EXCAVATOR/LOADER"),
    (34, "MISD", "MISD HARVESTER"),
    (35, "MISD", "MISD OTHERS"),
]


def _groups(ws, act_row, label_row, max_col):
    """Columns where ACT appears + the group label above-left of each."""
    out = []
    for c in range(1, max_col + 1):
        if str(ws.cell(act_row, c).value).strip() == "ACT":
            lbl = None
            for lc in range(c, max(1, c - 4), -1):
                v = ws.cell(label_row, lc).value
                if v is not None and str(v).strip():
                    lbl = str(v).strip()
                    break
            if lbl:
                out.append((c, lbl))
    return out


def extract_rows(src=None):
    wb = openpyxl.load_workbook(src or _file(), data_only=True)
    rate_rows, elig_rows, warnings = [], [], []
    seq = {"TW": 0, "PC": 0, "CV": 0, "ELIG": 0}

    def nid(k):
        seq[k] += 1
        return f"CHOLA-{k}-{seq[k]:04d}"

    def emit(k, **kw):
        kw.setdefault("insurer", INSURER)
        kw.setdefault("rule_type", "RATE")
        kw.setdefault("effect", "RATE")
        kw.setdefault("applies_on", "NET")
        kw.setdefault("confidence", 1.0)
        kw.setdefault("review_status", "PENDING")
        rate_rows.append(row(catalog_id=nid(k), **kw))

    # ------------------------------------------------------------ TW
    # Drive off the col-B merged ranges (each = one state block) and read the
    # per-row Type from col C, so variable-height blocks (GA has an extra SOD
    # row) are handled and no state is dropped.
    ws = wb[SRC_TW]
    n_zero = 0
    blocks = sorted((m.min_row, m.max_row) for m in ws.merged_cells.ranges
                    if m.min_col == 2 and m.min_row >= 6)
    # fall back to fixed blocks only if the sheet somehow has no merges
    if not blocks:
        blocks = [(r, r + 4) for r in range(6, 156, 6)]
    for r0, r1 in blocks:
        state = ws.cell(r0, 2).value
        if not (state and str(state).strip()):
            continue
        geo = _geo(state)
        for r in range(r0, r1 + 1):
            pol = _tw_policy(ws.cell(r, 3).value)
            if not pol:
                continue
            # combustion make grid
            for c0, mk, bands in TW_MAKES:
                for j, (lbl, sub, lo, hi) in enumerate(bands):
                    c = c0 + j
                    val = pct(ws.cell(r, c).value)
                    if val is None:
                        continue
                    if val == 0:
                        n_zero += 1
                    emit("TW", source_rule_id=f"{SRC_TW}!{get_column_letter(c)}{r}",
                         pay_in_pct=val, category="TW", sub_segment=sub, make=mk,
                         policy_type=pol,
                         cc_min=("" if lo is None else lo), cc_max=("" if hi is None else hi),
                         **geo, source_sheet=SRC_TW,
                         source_cell=f"{get_column_letter(c)}{r}",
                         source_text=f"{mk} {lbl} {pol}")
            # electric section (EV Type label is in col AQ=43, same rows)
            ev_pol = _tw_policy(ws.cell(r, 43).value) or pol
            for c, sub, lo, hi in TW_EV:
                val = pct(ws.cell(r, c).value)
                if val is None:
                    continue
                if val == 0:
                    n_zero += 1
                emit("TW", source_rule_id=f"{SRC_TW}!{get_column_letter(c)}{r}",
                     pay_in_pct=val, category="TW", sub_segment=sub, make="ALL ELECTRIC MAKES",
                     policy_type=ev_pol, **geo, source_sheet=SRC_TW,
                     source_cell=f"{get_column_letter(c)}{r}",
                     source_text=f"EV {sub} {ev_pol}")
    if n_zero:
        warnings.append({"scope": "sheet", "cell": SRC_TW,
                         "issue": f"{n_zero} TW cells carry a 0% pay-in (kept as RATE 0); "
                                  "confirm 0 means zero payout vs not offered"})
    warnings.append({"scope": "sheet", "cell": f"{SRC_TW}!AX:BE",
                     "issue": "'Hero vs Bajaj' differential and the standalone 'Bajaj' OEM-program "
                              "columns were NOT extracted (program-specific, needs clarification)"})

    # ------------------------------------------------------------ Private Car
    ws = wb[SRC_PC]
    groups = _groups(ws, act_row=6, label_row=4, max_col=66)
    for r, pol, sub, mk, lo, hi in _pc_rows():
        for act_c, lbl in groups:
            for c, p in ((act_c, "SATP"), (act_c + 1, pol)):
                # the ACT column is liability-only; the PACK/SOD column carries `pol`
                val = pct(ws.cell(r, c).value)
                if val is None:
                    continue
                if p == "SATP" and pol == "SAOD":
                    continue  # no ACT variant for SAOD rows beyond the PACK ones
                emit("PC", source_rule_id=f"{SRC_PC}!{get_column_letter(c)}{r}",
                     pay_in_pct=val, category="PVT_CAR", sub_segment=sub, make=mk,
                     policy_type=p,
                     applies_on=("OD" if p in ("COMP", "SAOD") else "NET"),
                     cc_min=("" if lo is None else lo), cc_max=("" if hi is None else hi),
                     **_geo(lbl), source_sheet=SRC_PC,
                     source_cell=f"{get_column_letter(c)}{r}",
                     source_text=f"{sub} {p} @ {lbl}")

    # ------------------------------------------------------------ CV
    ws = wb[SRC_CV]
    groups = _groups(ws, act_row=5, label_row=3, max_col=72)

    def _clean(v):  # collapse the newline-wrapped make/model cells
        return re.sub(r"\s+", " ", str(v).replace("\n", " ").strip()) if v else ""

    for r, cat, sub in CV_ROWS:
        mk = _clean(ws.cell(r, 5).value)     # col E = MAKE (as written in the grid)
        model = _clean(ws.cell(r, 6).value)  # col F = MODEL list
        for act_c, lbl in groups:
            for c, p in ((act_c, "SATP"), (act_c + 1, "COMP")):
                val = pct(ws.cell(r, c).value)
                if val is None:
                    continue
                emit("CV", source_rule_id=f"{SRC_CV}!{get_column_letter(c)}{r}",
                     pay_in_pct=val, category=cat, sub_segment=sub, make=mk, model=model,
                     policy_type=p, **_geo(lbl), source_sheet=SRC_CV,
                     source_cell=f"{get_column_letter(c)}{r}",
                     source_text=f"{sub} {p} @ {lbl}" + (f" [{mk}]" if mk else ""))

    # ------------------------------------------------------------ conditions -> eligibility
    def decl(cat, sub, pol, reason, cell, text):
        elig_rows.append(row(
            catalog_id=nid("ELIG"), source_rule_id=f"conditions!{cell}",
            insurer=INSURER, rule_type="ELIGIBILITY", effect="DECLINE",
            category=cat, sub_segment=sub, policy_type=pol,
            geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
            reason=reason, source_sheet="conditions", source_cell=cell,
            source_text=text, confidence=1.0, review_status="PENDING"))

    decl("TW", "", "SAOD",
         "TW SAOD has no payout and is blocked in system (conditions sheet)",
         "A15", "SAOD TW wont have any payouts,it will be blocked in system")
    # GCCV EV >3.5T and passenger-carrying EV (except eAUTO) declined PAN India.
    # This is conditional on FUEL=Electric AND weight >3.5T, dimensions we can't
    # bind in the atomic schema — so emitting it as a blanket PAN_INDIA decline
    # would wrongly block ALL GCV/PCV. Surface as a warning for manual handling.
    warnings.append({"scope": "conditions", "cell": "A9",
                     "issue": "GCCV EV exceeding 3.5T & passenger-carrying EV (except eAUTO) is "
                              "declined — conditional on fuel=Electric AND weight>3.5T (not "
                              "expressible atomically); enforce in platform / confirm handling"})

    return rate_rows, elig_rows, warnings


def geo_maps(src=None):
    """rto2clusters + cluster2states for the resolver (Chola zones via 'RTO ' sheet)."""
    wb = openpyxl.load_workbook(src or _file(), data_only=True)
    ws = wb[SRC_RTO]
    zone_label = {  # UW zone -> the grid column label that prices it
        "MUMBAI": "Mumbai/GA/Pune/Central MH", "PUNE": "Mumbai/GA/Pune/Central MH",
        "CENTRAL MH": "Mumbai/GA/Pune/Central MH", "CHENNAI": "TN-Chennai",
    }
    rto2clusters = {}
    for r in range(2, ws.max_row + 1):
        code, zone = ws.cell(r, 1).value, ws.cell(r, 3).value
        if not code or str(code).strip().upper() == "RTO_CODE":
            continue
        lbl = zone_label.get(str(zone).strip().upper())
        if lbl:
            rto2clusters.setdefault(str(code).strip().upper(), set()).add(lbl)
    # every grid label -> its states (so state queries match group columns)
    labels = set()
    wbm = wb
    for sheet, lr, ar in ((SRC_PC, 4, 6), (SRC_CV, 3, 5)):
        s = wbm[sheet]
        for c, lbl in _groups(s, act_row=ar, label_row=lr, max_col=s.max_column):
            labels.add(re.sub(r"\s+", " ", lbl.strip()))
    s = wbm[SRC_TW]
    for r0 in range(6, 156, 6):
        v = s.cell(r0, 2).value
        if v and str(v).strip():
            labels.add(re.sub(r"\s+", " ", str(v).strip()))
    cluster2states = {lbl: sorted(label_states(lbl)) for lbl in labels}
    cluster2states = {k: v for k, v in cluster2states.items() if v}
    return ({c: sorted(s) for c, s in rto2clusters.items()}, cluster2states)


if __name__ == "__main__":
    rates, eligs, warns = extract_rows()
    from collections import Counter
    print(f"CHOLA: {len(rates)} RATE, {len(eligs)} ELIG, {len(warns)} warnings")
    print("  by category:", dict(Counter(r["category"] for r in rates)))
    print("  by policy  :", dict(Counter(r["policy_type"] for r in rates)))
    r2c, c2s = geo_maps()
    print(f"  geo: {len(r2c)} rto codes, {len(c2s)} group labels")
    unresolved = [k for k, v in c2s.items() if not v]
    for s in rates[:3]:
        print("  e.g.", s["catalog_id"], s["sub_segment"], s["make"], s["policy_type"],
              "=", s["pay_in_pct"], "@", s["geo_label"], s["source_cell"])

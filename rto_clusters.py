"""
RTO-code -> commission-cluster mapping for SBI General, from the insurer-shared
"Broker Enabler GCV" workbook (sheet 'RTO Cluster': RTO Code | RTO Location |
RTO Code | RTO Cluster Name). Build-time only; build_deploy_data.py serializes
the result to api/_data/clusters.json for the serverless runtime.

Gives the resolution layer two joins the grid alone cannot provide:
  RTO2CLUSTER    "MH01" -> "MH - M"      (rto-code-level quoting)
  CLUSTER2STATES "MH - M" -> {"MAHARASHTRA"}   (state-level matching for
                 cluster-keyed rows, incl. all 1,170 highend rows)

The insurer file maps 5 codes to two clusters each (PB17, PB35, RJ38, TS07,
TS08); we keep the FIRST occurrence and surface each conflict as a warning.
"""
import glob

import openpyxl

import geo_normalize as gn

_FILE = glob.glob("rto_cluster_map_sbi.xlsx")[0]


def _prefix_majority():
    """code-prefix -> state, only where the master is unambiguous (not AP/BR/UP)."""
    from collections import defaultdict
    pre = defaultdict(set)
    for code, st in gn._CODE2STATE.items():
        pre[code[:2]].add(st)
    return {p: next(iter(s)) for p, s in pre.items() if len(s) == 1}


def load():
    ws = openpyxl.load_workbook(_FILE, data_only=True)["RTO Cluster"]
    prefix2state = _prefix_majority()
    rto2cluster, cluster2states, warnings = {}, {}, []
    for r in range(2, ws.max_row + 1):
        code, cluster = ws.cell(r, 3).value, ws.cell(r, 4).value
        if not (code and cluster):
            continue
        code, cluster = str(code).strip().upper(), str(cluster).strip()
        if code in rto2cluster:
            if rto2cluster[code] != cluster:
                warnings.append({
                    "scope": "rto_cluster_map", "cell": f"C{r}",
                    "issue": f"RTO {code} mapped to both '{rto2cluster[code]}' and "
                             f"'{cluster}' in the insurer file; kept the first",
                })
            continue
        rto2cluster[code] = cluster
        st = gn._CODE2STATE.get(code) or prefix2state.get(code[:2])
        if st:
            cluster2states.setdefault(cluster, set()).add(st)
    return rto2cluster, {c: sorted(s) for c, s in cluster2states.items()}, warnings


RTO2CLUSTER, CLUSTER2STATES, WARNINGS = load()


def cluster_state(cluster):
    """Single canonical state for a cluster, or '' if it spans several."""
    sts = CLUSTER2STATES.get(cluster, [])
    return sts[0] if len(sts) == 1 else ""


if __name__ == "__main__":
    multi = {c: s for c, s in CLUSTER2STATES.items() if len(s) > 1}
    print(f"{len(RTO2CLUSTER)} RTO codes -> {len(CLUSTER2STATES)} clusters, "
          f"{len(WARNINGS)} conflicts, {len(multi)} multi-state clusters")
    for c, s in sorted(multi.items()):
        print(f"  {c:14} -> {s}")

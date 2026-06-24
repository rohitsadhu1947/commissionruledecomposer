"""
Geography normalization layer, built from the RTO Cluster Master (RTO.xlsx).

The commission grids label geography in a mix of vocabularies:
  - real cities          : "Mumbai", "Pune", "Bangalore", "Kolkata"
  - region rollups       : "RO Maharashtra", "ROKarnataka", "Rest of West Bengal"
  - state-with-qualifier : "UTTAR PRADESH (Eastern)"
  - bare states          : "GOA", "BIHAR"
  - RTO codes (in notes) : "TN33", "GJ01"

Risk objects (and the notes' eligibility rules) speak in canonical UPPERCASE
state names. This module resolves any of the above to a canonical state so the
resolution engine can match a state-level query against city-level grid rows.

Source of truth: RTO.xlsx = a pasted API response (data.rtoData.allRtos[]).
Each RTO carries rtoName (code, e.g. "MH01") and rtoLongName
("MH01-Mumbai Central(Maharashtra)") with the state in the trailing parens.
"""
import json
import re
import openpyxl

_RTO_FILE = "RTO.xlsx"
_STATE_IN_PARENS = re.compile(r"\(([^()]*)\)[^()]*$")
_BAREWORD_KEY = re.compile(r"([{,\n]\s*)([A-Za-z_][A-Za-z0-9_]*)(\s*):")


def _load_all_rtos(path=_RTO_FILE):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    lines = [(str(r[0]) if r[0] is not None else "") for r in ws.iter_rows(values_only=True)]
    # the cells hold a JS-object literal (unquoted keys); quote keys -> valid JSON
    text = _BAREWORD_KEY.sub(r'\1"\2":', "\n".join(lines))
    return json.loads(text)["data"]["rtoData"]["allRtos"]


def _state_from_longname(longname):
    if not longname:
        return None
    m = _STATE_IN_PARENS.search(longname)
    return m.group(1).strip() if m else None


def _build_index(path=_RTO_FILE):
    code2state, city2state, states = {}, {}, set()
    for r in _load_all_rtos(path):
        st = _state_from_longname(r.get("rtoLongName"))
        if st:
            states.add(st.upper())
        if r.get("rtoName") and st:
            code2state[r["rtoName"].strip().upper()] = st.upper()
        if r.get("cityName") and st:
            city2state.setdefault(r["cityName"].strip().lower(), st.upper())
    return code2state, city2state, states


_CODE2STATE, _CITY2STATE, _STATES = _build_index()

# grid/notes spellings that differ from the RTO master's spelling
# (master canon: ORISSA, CHATTISGARH, PONDICHERRY, UTTARAKHAND, JAMMU AND KASHMIR)
_ALIASES = {
    "NCR": "DELHI",
    "ODISHA": "ORISSA",
    "CHHATTISGARH": "CHATTISGARH",
    "PUDUCHERRY": "PONDICHERRY",
    "UTTARANCHAL": "UTTARAKHAND",
    "UA/UK": "UTTARAKHAND",
    "J&K": "JAMMU AND KASHMIR",
    "J & K": "JAMMU AND KASHMIR",
    "JK": "JAMMU AND KASHMIR",
    "ANDAMANS": "ANDAMAN & NICOBAR ISLANDS",
    "ANDAMAN": "ANDAMAN & NICOBAR ISLANDS",
    "ANDAMAN & NICOBAR": "ANDAMAN & NICOBAR ISLANDS",
    # city spellings missing from / differing in the RTO master's city index
    "BARODA": "GUJARAT", "VADODARA": "GUJARAT",
    "KOCHI": "KERALA", "COCHIN": "KERALA",
    "NASIK": "MAHARASHTRA", "NASHIK": "MAHARASHTRA",
    "VIJAYWADA": "ANDHRA PRADESH", "VIJAYAWADA": "ANDHRA PRADESH",
    "VISHAKAPATTNAM": "ANDHRA PRADESH", "VISAKHAPATNAM": "ANDHRA PRADESH", "VIZAG": "ANDHRA PRADESH",
}


def canonical_states():
    return sorted(_STATES)


def to_state(label):
    """Resolve any geo label (city / region rollup / qualified state / RTO code /
    bare state) to a canonical UPPERCASE state name, or None if unresolvable."""
    if not label:
        return None
    raw = label.strip()
    up = raw.upper()

    # 1) exact RTO code (e.g. "TN33")
    if up in _CODE2STATE:
        return _CODE2STATE[up]

    # 2) exact city name (e.g. "Mumbai" -> MAHARASHTRA)
    if raw.lower() in _CITY2STATE:
        return _CITY2STATE[raw.lower()]

    # 3) already a known canonical state
    if up in _STATES:
        return up

    # 4) spelling aliases (incl. city aliases like BARODA->GUJARAT, J&K->...)
    if up in _ALIASES and _ALIASES[up] in _STATES:
        return _ALIASES[up]

    # 5) composite city/state list ("Ahmedabad, Baroda & Surat", "AS/ML/TR") ->
    #    resolve the parts FIRST (before the single-embedded-state fallback) so a
    #    multi-state label whose parts DISAGREE returns None instead of silently
    #    picking the first embedded state name.
    if any(sep in raw for sep in (",", "&", "/")):
        parts = [p.strip() for p in re.split(r"[,&/]", raw) if p.strip()]
        hits = {to_state(p) for p in parts}
        hits.discard(None)
        if len(hits) == 1:
            return next(iter(hits))
        if len(hits) > 1:
            return None  # genuinely multi-state -> let cluster2states handle it

    # 6) embedded state name (e.g. "RO Maharashtra", "UTTAR PRADESH (Eastern)",
    #    "Rest of West Bengal", "ROKarnataka"). Longest match wins.
    squashed = re.sub(r"[^A-Z]", "", up)
    hit = None
    for s in _STATES:
        if s in up or re.sub(r"[^A-Z]", "", s) in squashed:
            if hit is None or len(s) > len(hit):
                hit = s
    if hit:
        return hit

    # 7) '&' <-> 'AND' spelling drift ("JAMMU AND KASHMIR" vs "DAMAN & DIU")
    for swapped in (up.replace(" & ", " AND "), up.replace(" AND ", " & ")):
        if swapped != up and swapped in _STATES:
            return swapped

    # 8) strip a trailing qualifier in parens ("J&K (Non Preferred)",
    #    "TS Open (Others)") and retry on the bare label
    stripped = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
    if stripped and stripped != raw:
        return to_state(stripped)

    return None


if __name__ == "__main__":
    print(f"Loaded {len(_CODE2STATE)} RTO codes, {len(_CITY2STATE)} cities, {len(_STATES)} states")
    tests = ["Mumbai", "Navi Mumbai", "Pune", "RO Maharashtra", "Bangalore",
             "ROKarnataka", "Kolkata", "Rest of West Bengal", "NCR",
             "UTTAR PRADESH (Eastern)", "GOA", "TN33", "MH01", "GJ01",
             "HIMACHAL PRADESH", "MAHARASHTRA", "Nowhere City"]
    for t in tests:
        print(f"  {t!r:28} -> {to_state(t)!r}")

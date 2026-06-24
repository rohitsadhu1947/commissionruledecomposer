"""
Extractors for the four "lighter" insurer grids (4W-focused, partial coverage):

  TATA_AIG      Tata_Pvt car _Energise Broker.xlsx     flat table; PvtCar Package/
                SAOD x BusinessType x Fuel x NCB; pay already in % (OD basis)
  CAT_B         Motor Grid- CAT B Feb.xlsx             *insurer NOT named in the
                file* — registered as CAT_B pending team confirmation. PvtCar
                COMP/SAOD x Petrol/NonPetrol x NCB/NNCB priced by Preferred /
                Non-Preferred location tier per state; School Bus PCV; PAN-India
                sourcing restrictions for TW / GCV / PCV(3W & other) as declines.
  HDFC_ERGO     HDFC_SATP Grid NEW.xlsx                PvtCar SATP BDE% by
                location x fuel + its own RTO master (WB-01 style codes)
  ICICI_LOMBARD ICICI_Pvt Car Grid_June 26 Final M2B   PvtCar June'26 block:
                NB OD%, SAOD/COMP x fuel (NCB), Act-only, Non-NCB renewal, Used
                car; by state (EMG) / city (PMG). May'26 comparison block skipped.

All emit the shared 27-column atomic rows. Values that arrive as fractions are
x100; values already in percent are kept as-is (pct() handles both).
"""
import glob
import re

import openpyxl
from openpyxl.utils import get_column_letter

from catalog_schema import row
from geo_lookup import to_state

def _f(pat):
    return glob.glob(pat)[0]


def pct(v):
    if v is None or isinstance(v, str):
        return None
    v = float(v)
    return round(v * 100, 3) if -1.0 <= v <= 1.0 else round(v, 3)


def _mk(seq, ins, k):
    seq[k] = seq.get(k, 0) + 1
    return f"{ins.split('_')[0]}-{k}-{seq[k]:04d}"


def _base(ins, applies_on="NET"):
    return dict(insurer=ins, rule_type="RATE", effect="RATE", applies_on=applies_on,
                confidence=1.0, review_status="PENDING")


# ------------------------------------------------------------------ TATA AIG
def extract_tata(src=None):
    ws = openpyxl.load_workbook(src or _f("Tata_Pvt car*.xlsx"), data_only=True)["Sheet1"]
    rates, eligs, warns, seq = [], [], [], {}
    POL = {"Package": "COMP", "SAOD": "SAOD"}
    for r in range(2, ws.max_row + 1):
        if not ws.cell(r, 5).value:
            continue
        btype = str(ws.cell(r, 6).value or "").strip()      # Brand New/Renewal/Rollover
        section = str(ws.cell(r, 7).value or "").strip()    # Package/SAOD
        fuel = str(ws.cell(r, 9).value or "All").strip()
        ncb = str(ws.cell(r, 11).value or "All").strip()
        val = pct(ws.cell(r, 15).value)
        if val is None:
            continue
        pol = POL.get(section, section) + "-" + btype.upper().replace(" ", "")
        sub = "PVT CAR" + (f" [{fuel.upper()}]" if fuel != "All" else "")
        if ncb != "All":
            sub += " [NCB]" if ncb == "Yes" else " [NO NCB]"
        # TATA-1: discount-limit (L/M) + Add-On (N) columns. For this file every
        # row is 0/100/All, but flag if a re-upload ever varies them (silent collapse).
        dlo, dhi, addon = ws.cell(r, 12).value, ws.cell(r, 13).value, ws.cell(r, 14).value
        if (dlo not in (0, None) or dhi not in (100, None)
                or (addon and str(addon).strip().lower() != "all")):
            warns.append({"scope": "Sheet1", "cell": f"L{r}/M{r}/N{r}",
                          "issue": f"Discount limits/Add-On vary (L={dlo},M={dhi},N={addon}) but "
                                   "are not modelled — fold into sub_segment if they affect payout."})
        rates.append(row(
            catalog_id=_mk(seq, "TATA_AIG", "PC"), source_rule_id=f"Sheet1!O{r}",
            **_base("TATA_AIG", "OD"), pay_in_pct=val,
            category="PVT_CAR", sub_segment=sub, policy_type=pol,
            geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
            source_sheet="Sheet1", source_cell=f"O{r}",
            source_text=f"{btype} {section} {fuel} NCB={ncb}"))
    return rates, eligs, warns


# ------------------------------------------------------------------ CAT B
def extract_catb(src=None):
    wb = openpyxl.load_workbook(src or _f("Motor Grid- CAT B*.xlsx"), data_only=True)
    ws = wb["Pvt Comp "]
    rates, eligs, warns, seq = [], [], [], {}
    warns.append({"scope": "workbook", "cell": "-",
                  "issue": "Insurer name does not appear anywhere in 'Motor Grid- CAT B Feb.xlsx' "
                           "— registered as CAT_B; CONFIRM the actual insurer with the team"})
    # value grid: (policy, fuel) -> {(tier, ncb): col}
    GRID = {("COMP", "NON-PETROL"): {("Preferred", "NCB"): (8, 9), ("Non Preferred", "NCB"): (10, 9)},
            ("COMP", "PETROL"):     {("Preferred", "NCB"): (8, 10), ("Non Preferred", "NCB"): (10, 10)},
            ("SAOD", "NON-PETROL"): {("Preferred", "NCB"): (14, 9), ("Non Preferred", "NCB"): (16, 9)},
            ("SAOD", "PETROL"):     {("Preferred", "NCB"): (14, 10), ("Non Preferred", "NCB"): (16, 10)}}
    # state -> [(tier, location)]
    geo_rows, state = [], None
    for r in range(3, 39):
        s = ws.cell(r, 3).value
        if s and str(s).strip():
            state = str(s).strip()
        for c, tier in ((4, "Preferred"), (5, "Non Preferred")):
            loc = ws.cell(r, c).value
            if loc and str(loc).strip() and state:
                geo_rows.append((state, tier, str(loc).strip(), r))
    for (pol, fuel), cells in GRID.items():
        for (tier, _), (vc, vr) in cells.items():
            for ncb_off, ncb in ((0, "NCB"), (1, "NO NCB")):
                val = pct(ws.cell(vr, vc + ncb_off).value)
                if val is None:
                    continue
                cell = f"{get_column_letter(vc + ncb_off)}{vr}"
                for st, t, loc, gr in geo_rows:
                    if t != tier:
                        continue
                    rates.append(row(
                        catalog_id=_mk(seq, "CAT_B", "PC"), source_rule_id=f"Pvt Comp !{cell}",
                        **_base("CAT_B"), pay_in_pct=val,
                        category="PVT_CAR", sub_segment=f"PVT CAR [{fuel}] [{ncb}]",
                        policy_type=pol, geo_kind="RTO_CLUSTER",
                        geo_label=f"{loc} ({tier})",
                        canonical_state=to_state(loc) or to_state(st) or "",
                        source_sheet="Pvt Comp ", source_cell=cell,
                        source_text=f"{fuel} {ncb} {tier}: {st} / {loc} (geo row {gr})"))
    # School bus (hidden sheet)
    sb = wb["PCV Comp School Bus"]
    v = pct(sb.cell(3, 3).value)
    if v is not None:
        rates.append(row(
            catalog_id=_mk(seq, "CAT_B", "PCV"), source_rule_id="PCV Comp School Bus!C3",
            **_base("CAT_B"), pay_in_pct=v, category="PCV",
            sub_segment="SCHOOL BUS", policy_type="COMP",
            geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
            source_sheet="PCV Comp School Bus", source_cell="C3",
            source_text="School Bus, discount up to 99%, valid across all states (incl MISP)"))
    # PAN-India sourcing restrictions (IMP sheet)
    for cat, sub, why, cell in (
            ("TW", "", "Sourcing of TW bikes (Comp+SATP) restricted PAN India (0% commission)", "A9"),
            ("GCV", "", "Sourcing of GCV (Comp+SATP) restricted PAN India (0% commission)", "A10"),
            ("PCV", "PCV 3W/Other (excl School Bus)",
             "Sourcing of PCV (3W/Other) restricted PAN India (0% commission)", "A8")):
        eligs.append(row(
            catalog_id=_mk(seq, "CAT_B", "ELIG"), source_rule_id=f"IMP!{cell}",
            insurer="CAT_B", rule_type="ELIGIBILITY", effect="DECLINE",
            category=cat, sub_segment=sub, geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
            reason=why, source_sheet="IMP", source_cell=cell, source_text=why,
            confidence=1.0, review_status="PENDING"))
    # CAT_B-4: IMP sheet references Pvt Car SATP and Tractor grids that are NOT in
    # this workbook (only Comp/SAOD Pvt Car + School Bus present).
    warns.append({"scope": "IMP", "cell": "A3/A5",
                  "issue": "IMP notes reference 'Pvt Car SATP' (A3) and 'Tractor' grid changes "
                           "(A5), but this workbook contains NO Pvt Car SATP grid and NO tractor "
                           "grid — those products are NOT priced here. Source the separate grids."})
    # CAT_B-3: M14/M15 notes say SAOD has no NCB/NNCB distinction, yet the grid lays
    # out separate NCB/NNCB SAOD cells.
    warns.append({"scope": "Pvt Comp ", "cell": "M14/M15",
                  "issue": "Grid note: SAOD has no NCB/NNCB distinction and Petrol NCB=NNCB. "
                           "Extractor emits separate [NCB]/[NO NCB] SAOD rows; confirm whether "
                           "they should collapse."})
    return rates, eligs, warns


def geo_catb(src=None):
    wb = openpyxl.load_workbook(src or _f("Motor Grid- CAT B*.xlsx"), data_only=True)
    ws = wb["Pvt Comp "]
    c2s, state = {}, None
    for r in range(3, 39):
        s = ws.cell(r, 3).value
        if s and str(s).strip():
            state = str(s).strip()
        for c, tier in ((4, "Preferred"), (5, "Non Preferred")):
            loc = ws.cell(r, c).value
            if loc and str(loc).strip() and state:
                st = to_state(str(loc).strip()) or to_state(state)
                if st:
                    c2s.setdefault(f"{str(loc).strip()} ({tier})", set()).add(st)
    return {}, {k: sorted(v) for k, v in c2s.items()}


# ------------------------------------------------------------------ HDFC ERGO
def extract_hdfc(src=None):
    ws = openpyxl.load_workbook(src or _f("HDFC_SATP Grid*.xlsx"), data_only=True)["Grid"]
    rates, eligs, warns, seq = [], [], [], {}
    for r in range(3, ws.max_row + 1):
        st, loc, fuel = ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value
        val = pct(ws.cell(r, 6).value)
        if not (st and loc and fuel) or val is None:
            continue
        loc = str(loc).strip().rstrip("*")
        fz = str(fuel).strip().upper()
        cstate = to_state(str(st)) or ""
        # Grid footnote (H6/H7): listed BDE% is for >=1000cc; for 0-1000cc it is 5%
        # less. Emit BOTH cc bands so a cc-qualified quote resolves correctly.
        for lo, hi, adj, tag in ((1000, None, 0.0, ">=1000cc"), (0, 999, -5.0, "0-1000cc")):
            pv = round(val + adj, 3)
            rates.append(row(
                catalog_id=_mk(seq, "HDFC_ERGO", "PC"), source_rule_id=f"Grid!F{r}",
                **_base("HDFC_ERGO"), pay_in_pct=pv,
                category="PVT_CAR", sub_segment=f"PVT CAR SATP [{fz}]",
                policy_type="SATP", geo_kind="RTO_CLUSTER", geo_label=loc,
                canonical_state=cstate, cc_min=lo, cc_max=("" if hi is None else hi),
                source_sheet="Grid", source_cell=f"F{r}",
                source_text=f"{st} / {loc} {fuel} BDE% ({tag}{'; grid -5%' if adj else ''})"))
    return rates, eligs, warns


def geo_hdfc(src=None):
    ws = openpyxl.load_workbook(src or _f("HDFC_SATP Grid*.xlsx"), data_only=True)["Pvt Car SATP- RTO master"]
    r2c, c2s = {}, {}
    for r in range(2, ws.max_row + 1):
        st, loc, code = ws.cell(r, 1).value, ws.cell(r, 2).value, ws.cell(r, 4).value
        if not (loc and code):
            continue
        code = str(code).strip().upper().replace("-", "")
        loc = str(loc).strip()
        r2c.setdefault(code, set()).add(loc)
        s = to_state(str(st)) if st else None
        if s:
            c2s.setdefault(loc, set()).add(s)
    return ({k: sorted(v) for k, v in r2c.items()}, {k: sorted(v) for k, v in c2s.items()})


# ------------------------------------------------------------------ ICICI Lombard
def extract_icici(src=None):
    ws = openpyxl.load_workbook(src or _f("ICICI_Pvt Car Grid*.xlsx"), data_only=True)["Sheet1"]
    rates, eligs, warns, seq = [], [], [], {}
    # June'26 block, cols C..N. (May'26 comparison block P..AC intentionally skipped.)
    COLS = [
        (3,  "COMP-NB (1+3/3+3)", "", "OD"),
        (4,  "SAOD", "[PETROL] [NCB]", "OD"), (5, "SAOD", "[CNG] [NCB]", "OD"),
        (6,  "SAOD", "[DIESEL] [NCB]", "OD"), (7, "SAOD", "[ELECTRIC] [NCB]", "OD"),
        (8,  "COMP", "[PETROL] [NCB]", "OD"), (9, "COMP", "[CNG] [NCB]", "OD"),
        (10, "COMP", "[DIESEL] [NCB]", "OD"), (11, "COMP", "[ELECTRIC] [NCB]", "OD"),
        (12, "SATP", "", "NET"),
        (13, "SAOD/COMP [NON-NCB RENEWAL]", "", "OD"),
        (14, "COMP", "USED CAR", "OD"),
    ]
    n_zero = 0
    for r in range(4, ws.max_row + 1):
        tier, geo = ws.cell(r, 1).value, ws.cell(r, 2).value
        if not (tier and geo):
            continue
        geo = str(geo).strip()
        for c, pol, qual, basis in COLS:
            val = pct(ws.cell(r, c).value)
            if val is None:
                continue
            if val == 0:
                n_zero += 1
            sub = ("PC USED CAR" if qual == "USED CAR"
                   else "PVT CAR" + (f" {qual}" if qual else ""))
            rates.append(row(
                catalog_id=_mk(seq, "ICICI_LOMBARD", "PC"),
                source_rule_id=f"Sheet1!{get_column_letter(c)}{r}",
                **_base("ICICI_LOMBARD", basis), pay_in_pct=val,
                category="PVT_CAR", sub_segment=sub, policy_type=pol,
                geo_kind="STATE_OR_CITY", geo_label=geo,
                canonical_state=to_state(geo) or "",
                source_sheet="Sheet1", source_cell=f"{get_column_letter(c)}{r}",
                source_text=f"{tier} {geo} {pol} {qual}".strip()))
    if n_zero:
        warns.append({"scope": "Sheet1", "cell": "L col",
                      "issue": f"{n_zero} ICICI cells carry 0% (mostly Act-only) — kept as RATE 0; "
                               "confirm 0 means zero payout vs not offered"})
    warns.append({"scope": "Sheet1", "cell": "P:AC",
                  "issue": "ICICI May'26 comparison block (cols P-AC) intentionally not extracted "
                           "(June'26 is the active grid)"})
    return rates, eligs, warns


# ------------------------------------------------------------------ registry
INSURERS = [
    ("TATA_AIG", extract_tata, None),
    ("CAT_B", extract_catb, geo_catb),
    ("HDFC_ERGO", extract_hdfc, geo_hdfc),
    ("ICICI_LOMBARD", extract_icici, None),
]


if __name__ == "__main__":
    from collections import Counter
    for name, fn, geo in INSURERS:
        rates, eligs, warns = fn()
        cats = dict(Counter(r["category"] for r in rates))
        print(f"{name:14} {len(rates):>4} RATE  {len(eligs)} ELIG  {len(warns)} warn  {cats}")
        if geo:
            r2c, c2s = geo()
            print(f"{'':14} geo: {len(r2c)} rto codes, {len(c2s)} labels")
        for s in rates[:2]:
            print(f"{'':14} e.g. {s['catalog_id']} {s['sub_segment']} {s['policy_type']} "
                  f"= {s['pay_in_pct']} @ {s['geo_label']} ({s['source_cell']})")

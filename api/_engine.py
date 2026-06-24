"""
Runtime engine for the Vercel function. Loads precomputed JSON from _data/
(no Excel parsing at runtime) and provides: geo normalization, the quote
resolution engine, and in-memory approved-catalog export (csv/json/xlsx bytes).

Mirrors resolution_engine.py + export_approved.py but file-free and stateless.
"""
import csv
import io
import json
import os
import re

_DATA = os.path.join(os.path.dirname(__file__), "_data")


def _load(name):
    with open(os.path.join(_DATA, name)) as f:
        return json.load(f)


RATE_RULES = _load("rate_rules.json")
ELIG_RULES = _load("elig_rules.json")
CATALOG = _load("catalog.json")
_GEO = _load("geo.json")
_CODE2STATE = _GEO["code2state"]
_CITY2STATE = _GEO["city2state"]
_STATES = set(_GEO["states"])
_ALIASES = _GEO.get("aliases", {})

try:  # per-insurer RTO->cluster maps (rto_code quoting + state matching for cluster rows)
    _CLUSTERS = _load("clusters.json")
except FileNotFoundError:
    _CLUSTERS = {}
# per-insurer lookups: {insurer: {code: {labels}}} / {insurer: {label: {states}}}
_RTO2CLUSTERS = {ins: {k: set(v) for k, v in m.get("rto2clusters", {}).items()}
                 for ins, m in _CLUSTERS.items()}
_CLUSTER2STATES = {ins: {c: set(s) for c, s in m.get("cluster2states", {}).items()}
                   for ins, m in _CLUSTERS.items()}


def _clkey(label):
    """Punctuation/case-insensitive cluster key (GEO-07): 'Bad KA'/'BAD KA',
    'GOOD CG'/'Good CG' collapse to one key for state lookups."""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", str(label or ""))   # drop trailing qualifier
    s = re.sub(r"_+lowcd2.*$", "", s, flags=re.I)            # drop CD-band suffix
    return re.sub(r"[^a-z0-9]", "", s.lower())


# normalized cluster -> states, merged across case/punctuation variants
_CLUSTER2STATES_NORM = {}
for _ins, _m in _CLUSTER2STATES.items():
    d = _CLUSTER2STATES_NORM.setdefault(_ins, {})
    for _lbl, _sts in _m.items():
        d.setdefault(_clkey(_lbl), set()).update(_sts)

# multi-state region labels used by the grid that fold to no single state
_REGION2STATES = {
    "NORTH EAST": {"ASSAM", "ARUNACHAL PRADESH", "MANIPUR", "MEGHALAYA",
                   "MIZORAM", "NAGALAND", "SIKKIM", "TRIPURA"},
}

CATALOG_COLUMNS = [
    "catalog_id", "source_rule_id", "rule_type", "effect", "pay_in_pct", "applies_on",
    "insurer", "category", "sub_segment", "make", "model", "match",
    "cc_min", "cc_max", "policy_type", "geo_kind", "geo_label", "canonical_state", "rto_code",
    "age_min", "age_max", "reason", "source_sheet", "source_cell", "source_text",
    "confidence", "review_status",
]
APPROVED_COLUMNS = CATALOG_COLUMNS + ["reviewer", "reviewed_ts", "edited"]


# ----------------------------------------------------------------- geo
def to_state(label):
    if not label:
        return None
    raw = str(label).strip()
    up = raw.upper()
    if up in _CODE2STATE:
        return _CODE2STATE[up]
    if raw.lower() in _CITY2STATE:
        return _CITY2STATE[raw.lower()]
    if up in _STATES:
        return up
    if up in _ALIASES and _ALIASES[up] in _STATES:
        return _ALIASES[up]
    # composite list resolved BEFORE single-embedded-state fallback (GEO-01)
    if any(sep in raw for sep in (",", "&", "/")):
        hits = {to_state(p.strip()) for p in re.split(r"[,&/]", raw) if p.strip()}
        hits.discard(None)
        if len(hits) == 1:
            return next(iter(hits))
        if len(hits) > 1:
            return None
    squashed = re.sub(r"[^A-Z]", "", up)
    hit = None
    for s in _STATES:
        if s in up or re.sub(r"[^A-Z]", "", s) in squashed:
            if hit is None or len(s) > len(hit):
                hit = s
    if hit:
        return hit
    for swapped in (up.replace(" & ", " AND "), up.replace(" AND ", " & ")):
        if swapped != up and swapped in _STATES:
            return swapped
    stripped = re.sub(r"\s*\([^)]*\)\s*$", "", raw).strip()
    if stripped and stripped != raw:
        return to_state(stripped)
    return None


def _norm_state(s):
    if s is None:
        return None
    return to_state(s) or str(s).strip().upper()


# -------------------------------------------------------- resolution engine
def _cc_in_band(cc, band):
    if band is None:
        return True
    lo, hi = band.get("min_cc"), band.get("max_cc")
    if lo is not None and not (cc > lo):
        return False
    if hi is not None and not (cc <= hi):
        return False
    return True


def _scope_matches(scope, risk):
    bound = 0

    def check(dim, predicate):
        nonlocal bound
        if predicate is None:
            return True
        bound += 1
        return risk.get(dim) is not None and predicate(risk[dim])

    ok = True
    ok &= check("category", (lambda v: v == scope["category"]) if scope.get("category") else None)
    ok &= check("sub_segment", (lambda v: v == scope["sub_segment"]) if scope.get("sub_segment") else None)
    ok &= check("policy_type", (lambda v: v in scope["policy_type"]) if scope.get("policy_type") else None)
    ok &= check("rto_cluster",
                (lambda v: v.strip().lower() == scope["rto_cluster"].strip().lower()) if scope.get("rto_cluster") else None)
    ok &= check("state",
                (lambda v: _norm_state(v) == _norm_state(scope["state"])) if scope.get("state") else None)
    if scope.get("cc_band") is not None:
        bound += 1
        if risk.get("cc") is None or not _cc_in_band(risk["cc"], scope["cc_band"]):
            ok = False
    return ok, bound


def _as_list(v):
    return v if isinstance(v, list) else [v]


def _make_match(risk_make, target):
    if target == "*OTHER*":
        return True
    return risk_make is not None and target.lower() in risk_make.lower()


def _model_match(risk_model, models, mode=None):
    if models is None or "ALL" in models:
        return True
    if risk_model is None:
        return False
    rm = risk_model.lower()
    return any(m.lower() in rm or rm in m.lower() for m in models)


def check_eligibility(elig_rules, risk):
    reasons = []
    rstate = _norm_state(risk.get("state"))
    for r in elig_rules:
        sc, eff = r["scope"], r["effect"]
        if sc.get("category") and risk.get("category") != sc["category"]:
            continue
        if sc.get("sub_segment") and risk.get("sub_segment") not in _as_list(sc["sub_segment"]):
            continue
        if sc.get("policy_type") and risk.get("policy_type") not in _as_list(sc["policy_type"]):
            continue
        if eff.get("mode") == "ALLOW_ONLY":
            st_ok = rstate in {_norm_state(s) for s in eff.get("allowed_states", [])}
            rto_ok = risk.get("rto_code") in eff.get("allowed_rto_codes", [])
            if not (st_ok or rto_ok):
                reasons.append((r["rule_id"], eff["reason"]))
            continue
        if "age_years_min" in sc:
            if risk.get("age_years") is None or risk["age_years"] < sc["age_years_min"]:
                continue
        if "states" in sc and "make_models" not in sc and "make_state_declines" not in sc:
            if rstate not in {_norm_state(s) for s in sc["states"]}:
                continue
            reasons.append((r["rule_id"], eff["reason"]))
            continue
        if "make_models" in sc:
            if "states" in sc and rstate not in {_norm_state(s) for s in sc["states"]}:
                continue
            hit = any(_make_match(risk.get("make"), mm["make"]) and _model_match(risk.get("model"), mm.get("models"), mm.get("match"))
                      for mm in sc["make_models"])
            if hit:
                reasons.append((r["rule_id"], eff["reason"]))
            continue
        if "make_state_declines" in sc:
            for msd in sc["make_state_declines"]:
                if rstate in {_norm_state(s) for s in msd["states"]} and _make_match(risk.get("make"), msd["make"]):
                    reasons.append((r["rule_id"], f'{eff["reason"]} ({msd["make"]})'))
                    break
            continue
        if "age_years_min" in sc:
            reasons.append((r["rule_id"], eff["reason"]))
    return reasons


def resolve_rate(rules, risk):
    candidates = []
    for r in rules:
        if r["rule_type"] != "RATE":
            continue
        ok, spec = _scope_matches(r["scope"], risk)
        if ok:
            candidates.append((spec, r["precedence"], r))
    if not candidates:
        return None
    candidates.sort(key=lambda t: (t[0], t[1]), reverse=True)
    w = candidates[0][2]
    return {"pay_in_pct": w["effect"]["value"], "applies_on": w["effect"]["applies_on"],
            "rule_id": w["rule_id"], "trace": [c[2]["rule_id"] for c in candidates]}


def quote(risk):
    declines = check_eligibility(ELIG_RULES, risk)
    if declines:
        return {"decision": "DECLINE", "reasons": declines}
    base = resolve_rate(RATE_RULES, risk)
    if base is None:
        return {"decision": "NO_RATE", "reason": "No matching rate rule"}
    return {"decision": "ALLOW", "pay_in_pct": base["pay_in_pct"], "base_pct": base["pay_in_pct"],
            "applies_on": base["applies_on"], "rate_rule": base["rule_id"], "modifiers": []}


# ----------------------------------------- unified catalog resolver (all categories)
# Resolves a risk against the FLAT catalog rows (the deliverable the team loads),
# so the test console can exercise 4W + CV exactly as they will run in production.

def _in_band(val, lo, hi):
    if lo in ("", None) and hi in ("", None):
        return True
    if val is None:
        return False
    if lo not in ("", None) and val < float(lo):
        return False
    if hi not in ("", None) and val > float(hi):
        return False
    return True


def _geo_ok(r, risk):
    """Return (matched, specificity). rto_code/cluster is most specific, then state."""
    if r.get("geo_kind") == "PAN_INDIA":
        return True, 0
    ins = r.get("insurer") or ""
    gl = (str(r.get("geo_label") or "")).strip()
    cl = (str(risk.get("cluster") or "")).strip()
    rto = (str(risk.get("rto_code") or "")).strip().upper()
    if rto and not cl:
        # an RTO code pins the insurer's cluster set (per-product vocabularies
        # mean one code can map to several labels for the same insurer)
        labels = {_clkey(l) for l in _RTO2CLUSTERS.get(ins, {}).get(rto, ())}
        if not labels:
            return False, 0  # unknown RTO code for this insurer: fail loudly
        if _clkey(gl) in labels:
            return True, 2
        # fall through to state matching only if the rule is state-keyed
    elif cl and gl and _clkey(cl) == _clkey(gl):
        return True, 2
    rs = _norm_state(risk.get("state"))
    if not rs and rto:  # an RTO code implies its state for state-keyed rules
        rs = _CODE2STATE.get(rto)
    cs = (str(r.get("canonical_state") or "")).strip()
    if rs:
        if cs and _norm_state(cs) == rs:
            return True, 1
        if gl and _norm_state(gl) == rs:
            return True, 1
        # cluster/state-group row whose span covers the queried state
        # (normalized lookup folds case/punctuation variants — GEO-07)
        if rs in _CLUSTER2STATES_NORM.get(ins, {}).get(_clkey(gl), ()):
            return True, 1
        if rs in _REGION2STATES.get(gl.upper(), ()):
            return True, 1
        return False, 0
    if not cl and not rto:  # no geo supplied at all -> unconstrained
        return True, 0
    return False, 0


def _loose(a, b):
    a, b = (str(a or "")).strip().lower(), (str(b or "")).strip().lower()
    if not a or not b:
        return False
    return a == b or a in b or b in a


# (catch-all make tiers are now a pure wildcard handled in _make_ok, so the old
#  _tier_names/_build_tier_index/_TIER_IDX sibling machinery is no longer needed.)


_EXCL_RE = re.compile(r"\b(?:OTHER THAN|EXCEPT|EXCLUDING|EXCL\.?|NOT)\b", re.I)


def _split_makes(text):
    """A make-list string -> list of make tokens. Handles ',', '&', '/', 'and',
    trailing '*' footnote marks and parentheses."""
    s = re.sub(r"[()*]", " ", str(text or ""))
    parts = re.split(r"\s*(?:,|&|/|\band\b)\s*", s, flags=re.I)
    return [p.strip() for p in parts if p.strip()]


def _make_in_list(rm, text):
    return any(_loose(rm, tok) for tok in _split_makes(text))


def _make_ok(risk_make, rule_make):
    """Tier-aware make check. Returns (ok, score).
    - explicit exclusions ('Other than Tata', 'Non-Tata', 'All except Tata & AL',
      'All excluding Volvo and Scania') -> match everything NOT in the excluded list
    - generic catch-all ('All Other Make/Models', 'Other Makes', 'All Make') ->
      WILDCARD at zero specificity; scoring lets a named tier that matches win
    - otherwise substring/equality match on the named make(s)."""
    rm = (str(risk_make or "")).strip()
    tm = (str(rule_make or "")).strip()
    if not rm or not tm:
        return True, 0  # unconstrained
    up = tm.upper()
    m = _EXCL_RE.search(up)
    if m:
        excluded = up[m.end():]
        return (not _make_in_list(rm, excluded)), 1
    if up.startswith("NON-") or up.startswith("NON "):  # 'Non-Tata'
        return (not _loose(rm, up[4:].strip())), 1
    if "OTHER" in up or "ALL MAKE" in up or "ALL MODEL" in up or up in ("ALL", "ANY"):
        # catch-all tier — match anything at zero specificity.
        return True, 0
    # named tier, possibly a list ('Tata/Maruti/Mahindra'): match if risk is any token
    return (_loose(rm, tm) or _make_in_list(rm, tm)), 1


# ---- policy_type family matching (grids use compound labels) ----
def _pol_families(label):
    """Base policy families present in a (possibly compound) policy label."""
    u = str(label or "").upper()
    fams = set()
    if "SAOD" in u or "SOD" in u:
        fams.add("SAOD")
    if "SATP" in u or "ACT" in u or u.strip() in ("TP",) or " TP" in u:
        fams.add("SATP")
    if "COMP" in u or "PACKAGE" in u or "PACK" in u:
        fams.add("COMP")
    return fams


def _policy_ok(risk_pol, rule_pol):
    """Returns (ok, score). Exact label match scores highest; a clean family
    query ('COMP'/'SATP'/'SAOD') matches any compound label of that family."""
    if not risk_pol:
        return True, 0  # unconstrained
    if not rule_pol:
        return True, 0
    if str(risk_pol).strip().upper() == str(rule_pol).strip().upper():
        return True, 2  # exact label
    rfam = _pol_families(risk_pol)
    if rfam and (rfam & _pol_families(rule_pol)):
        return True, 1  # family match (e.g. 'COMP' vs 'COMP & SATP')
    return False, 0


def _rate_match(r, risk, tier_idx=None):  # tier_idx kept for call-site compat
    score = 0
    if risk.get("sub_segment"):
        if r.get("sub_segment") != risk["sub_segment"]:
            return False, 0
        score += 1
    if risk.get("policy_type"):
        pok, psc = _policy_ok(risk["policy_type"], r.get("policy_type"))
        if not pok:
            return False, 0
        score += psc
    if risk.get("make") and r.get("make"):
        mok, msc = _make_ok(risk["make"], r["make"])
        if not mok:
            return False, 0
        score += msc
    if risk.get("model") and r.get("model"):
        if not _loose(risk["model"], r["model"]):
            return False, 0
        score += 1
    # cc/age: a blank input is a wildcard (surfaces every band as a candidate);
    # a supplied value must fall inside the band or the rule is rejected.
    has_cc = r.get("cc_min", "") != "" or r.get("cc_max", "") != ""
    if risk.get("cc") is not None and has_cc:
        if not _in_band(risk["cc"], r.get("cc_min", ""), r.get("cc_max", "")):
            return False, 0
        score += 1
    has_age = r.get("age_min", "") != "" or r.get("age_max", "") != ""
    if risk.get("age_years") is not None and has_age:
        if not _in_band(risk["age_years"], r.get("age_min", ""), r.get("age_max", "")):
            return False, 0
        score += 1
    gok, gsc = _geo_ok(r, risk)
    if not gok:
        return False, 0
    return True, score + gsc


def _decline_match(r, risk):
    """ELIGIBILITY DECLINE row vs risk: every dimension the rule binds must match
    (sub_segment, policy, geo, age/cc, then make/model)."""
    if r.get("sub_segment") and risk.get("sub_segment") and r["sub_segment"] != risk["sub_segment"]:
        return False
    if r.get("policy_type"):
        # a policy-scoped decline fires only when the risk policy is the same family
        if not risk.get("policy_type"):
            return False
        pok, _ = _policy_ok(risk["policy_type"], r["policy_type"])
        if not pok:
            return False
    # F06: a GEO-scoped decline must NOT fire when the risk omits geo entirely
    # (absence of geo is not a wildcard hit for a state/cluster-specific block).
    geo_bound = (r.get("geo_kind") not in ("PAN_INDIA", "", "ELSEWHERE")
                 and (str(r.get("geo_label") or "").strip()
                      or str(r.get("canonical_state") or "").strip()))
    risk_has_geo = bool(risk.get("state") or risk.get("cluster") or risk.get("rto_code"))
    if geo_bound and not risk_has_geo:
        return False
    gok, _ = _geo_ok(r, risk)
    if not gok:
        return False
    # age/cc-gated declines fire only when the risk supplies a value in the band
    for lo_k, hi_k, rk in (("age_min", "age_max", "age_years"), ("cc_min", "cc_max", "cc")):
        if r.get(lo_k, "") != "" or r.get(hi_k, "") != "":
            if risk.get(rk) is None:
                return False
            if not _in_band(risk[rk], r.get(lo_k, ""), r.get(hi_k, "")):
                return False
    if not r.get("make"):
        # no make condition: fires on segment/policy/geo. A PAN_INDIA row with
        # nothing else bound is a category-wide sourcing restriction.
        return bool(r.get("policy_type") or r.get("sub_segment")
                    or r.get("geo_kind") == "PAN_INDIA")
    if not risk.get("make"):
        return False
    # F03: respect exclusion phrasing ('All excl. Bajaj and TVS') so the decline
    # does NOT fire for the excluded (allowed) makes.
    mok, _ = _make_ok(risk["make"], r["make"])
    if not mok:
        return False
    dm = (str(r.get("model") or "")).strip().lower()
    rm = (str(risk.get("model") or "")).strip().lower()
    if dm == "" or "all model" in dm or "all variants" in dm:
        return True
    if dm.startswith("all except"):
        return bool(rm) and (rm not in dm.split("except", 1)[1])
    if rm and rm in dm:
        return True
    return rm == ""  # make on the decline list, no model given -> flag for review


def _brief(r):
    return {k: r.get(k) for k in (
        "catalog_id", "pay_in_pct", "policy_type", "sub_segment", "make", "model",
        "geo_label", "canonical_state", "cc_min", "cc_max", "age_min", "age_max",
        "source_sheet", "source_cell")}


def _decl_brief(r):
    return {"catalog_id": r["catalog_id"], "insurer": r.get("insurer", ""),
            "reason": r.get("reason", ""), "source_sheet": r.get("source_sheet", ""),
            "source_cell": r.get("source_cell", ""), "source_text": r.get("source_text", "")}


def quote_catalog(risk):
    cat = risk.get("category")
    ins = risk.get("insurer")
    rules = [r for r in CATALOG["rules"] if r.get("category") == cat
             and (not ins or r.get("insurer") == ins)]
    matched_declines = [r for r in rules
                        if r.get("rule_type") == "ELIGIBILITY" and r.get("effect") == "DECLINE"
                        and _decline_match(r, risk)]
    cands = []
    for r in rules:
        if r.get("rule_type") != "RATE":
            continue
        ok, score = _rate_match(r, risk)
        if ok:
            cands.append((score, r))
    # F08: deterministic ranking — highest specificity, then most conservative
    # (lowest) pay-in on a tie, then catalog_id, so the winner is never arbitrary.
    cands.sort(key=lambda t: (-t[0], float(t[1].get("pay_in_pct") or 0), t[1].get("catalog_id", "")))

    best = cands[0][1] if cands else None
    # F05: a decline only blocks a rate from the SAME insurer (no cross-insurer
    # leak when the query omits `insurer`).
    win_ins = (best or {}).get("insurer") if best else ins
    blocking = [d for d in matched_declines if not win_ins or d.get("insurer") == win_ins]
    same_ins_declines = [_decl_brief(d) for d in (blocking if best else matched_declines)]

    # F01: a fired decline OVERRIDES an available rate (block beats pay-in).
    if blocking:
        return {"decision": "DECLINE", "reasons": same_ins_declines,
                "reason": same_ins_declines[0]["reason"] or "Declined for these attributes",
                "declines": same_ins_declines,
                "n_candidates": len(cands)}
    if not cands:
        decls = [_decl_brief(d) for d in matched_declines]
        return {"decision": "DECLINE" if decls else "NO_RATE",
                "reasons": decls, "reason": "No matching rate rule for these attributes",
                "declines": decls}
    return {"decision": "ALLOW", "pay_in_pct": best.get("pay_in_pct"),
            "base_pct": best.get("pay_in_pct"), "applies_on": best.get("applies_on", "NET"),
            "rate_rule": best["catalog_id"], "matched": _brief(best),
            "candidates": [_brief(c[1]) for c in cands[:8]],
            "n_candidates": len(cands), "declines": same_ins_declines, "modifiers": []}


def _build_options():
    """insurer -> category -> {geo, subs{policy, make, has_cc, has_age}}."""
    out = {}
    for r in CATALOG["rules"]:
        if r.get("rule_type") != "RATE":
            continue
        ins = r.get("insurer") or ""
        cat = r.get("category") or ""
        sub = r.get("sub_segment") or ""
        c = out.setdefault(ins, {}).setdefault(cat, {"subs": {}, "geo": set()})
        s = c["subs"].setdefault(sub, {"policy": set(), "make": set(),
                                       "has_cc": False, "has_age": False})
        if r.get("policy_type"):
            s["policy"].add(r["policy_type"])
        if r.get("make"):
            s["make"].add(r["make"])
        if r.get("cc_min", "") != "" or r.get("cc_max", "") != "":
            s["has_cc"] = True
        if r.get("age_min", "") != "" or r.get("age_max", "") != "":
            s["has_age"] = True
        if r.get("geo_label"):
            c["geo"].add(str(r["geo_label"]))
    res = {}
    for ins, cats in out.items():
        res[ins] = {}
        for cat, c in cats.items():
            res[ins][cat] = {"geo": sorted(c["geo"]), "subs": {}}
            for sub, s in c["subs"].items():
                res[ins][cat]["subs"][sub] = {
                    "policy": sorted(s["policy"]), "make": sorted(s["make"]),
                    "has_cc": s["has_cc"], "has_age": s["has_age"]}
    return res


OPTIONS = _build_options()

# ----------------------------------------------- monthly grid overrides
# Uploaded grids (parsed server-side, stored in KV) replace the baked rows for
# their insurer. The merged view is recomputed at most once per minute per
# warm instance, and immediately after a commit.
_BASE_CATALOG = CATALOG
_REFRESH = {"ts": 0.0}


def _assemble():
    import _store as store
    over = {}
    for ins in store.overridden_insurers():
        g = store.get_grid(ins)
        if g and g.get("rules"):
            over[ins] = g
    if not over:
        return _BASE_CATALOG
    rules = [r for r in _BASE_CATALOG["rules"] if r.get("insurer") not in over]
    meta = dict(_BASE_CATALOG["meta"])
    meta["overrides"] = {}
    # F4: rebuild warnings — drop baked warnings for overridden insurers, append
    # each override's own extraction warnings (the upload stores them in its meta).
    warns = [w for w in _BASE_CATALOG["meta"].get("warnings", [])
             if w.get("insurer") not in over]
    for ins, g in over.items():
        rules = rules + g["rules"]
        gm = g.get("meta", {})
        meta["overrides"][ins] = gm
        for w in gm.get("warnings", []):
            warns.append({**w, "insurer": ins})
    meta["warnings"] = warns
    from collections import Counter
    meta["counts"] = {
        "total": len(rules),
        "rate": sum(1 for r in rules if r.get("rule_type") == "RATE"),
        "eligibility": sum(1 for r in rules if r.get("rule_type") == "ELIGIBILITY"),
        "by_insurer": dict(Counter(r.get("insurer") for r in rules)),
        "by_category": dict(Counter(r.get("category") for r in rules)),
    }
    return {"meta": meta, "rules": rules}


def refresh(force=False):
    """Re-merge baked catalog + KV overrides; rebuild the derived indexes."""
    global CATALOG, OPTIONS, STATES_ALL
    import time
    now = time.time()
    if not force and now - _REFRESH["ts"] < 60:
        return
    _REFRESH["ts"] = now
    CATALOG = _assemble()
    OPTIONS = _build_options()
    STATES_ALL = _all_states()


def _all_states():
    """Every state reachable by any catalog rule (canonical folds + cluster spans)."""
    out = set()
    for r in CATALOG.get("rules", []):
        cs = (r.get("canonical_state") or "").strip()
        if cs in _STATES:
            out.add(cs)
        gl = (str(r.get("geo_label") or "")).strip()
        out.update(_CLUSTER2STATES.get(r.get("insurer") or "", {}).get(gl, ()))
        out.update(_REGION2STATES.get(gl.upper(), ()))
    return sorted(out)


STATES_ALL = _all_states()


# ------------------------------------------------- approved-catalog export
def apply_reviews(rules, review):
    out, counts = [], {"confirmed": 0, "rejected": 0, "pending": 0, "edited": 0}
    for r in rules:
        d = review.get(r["catalog_id"])
        row = dict(r)
        row.setdefault("reviewer", "")
        row.setdefault("reviewed_ts", "")
        row.setdefault("edited", "")
        if not d:
            row["review_status"] = "PENDING"
            counts["pending"] += 1
            out.append(row)
            continue
        status = d.get("status", "PENDING")
        if status == "REJECTED":
            counts["rejected"] += 1
            continue
        edited = False
        if d.get("pay_in_pct") not in (None, "") and str(d["pay_in_pct"]) != str(r.get("pay_in_pct")):
            row["pay_in_pct"] = d["pay_in_pct"]
            edited = True
        if d.get("reason") not in (None, "") and d["reason"] != r.get("reason"):
            row["reason"] = d["reason"]
            edited = True
        row["review_status"] = status
        row["reviewer"] = d.get("reviewer", "")
        row["reviewed_ts"] = d.get("ts", "")
        row["edited"] = "yes" if edited else ""
        counts["confirmed" if status == "CONFIRMED" else "pending"] += 1
        if edited:
            counts["edited"] += 1
        out.append(row)
    return out, counts


def export_csv_bytes(rows):
    buf = io.StringIO()
    w = csv.DictWriter(buf, fieldnames=APPROVED_COLUMNS)
    w.writeheader()
    for r in rows:
        w.writerow({c: r.get(c, "") for c in APPROVED_COLUMNS})
    return buf.getvalue().encode("utf-8")


def export_json_bytes(rows, meta):
    return json.dumps({"meta": meta, "rules": rows}, indent=2).encode("utf-8")


def export_xlsx_bytes(rows, meta):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Rules"
    hf, ht = PatternFill("solid", fgColor="1F2937"), Font(bold=True, color="FFFFFF")
    fills = {"RATE": PatternFill("solid", fgColor="E8F5E9"),
             "ALLOW": PatternFill("solid", fgColor="FFF8E1"),
             "DECLINE": PatternFill("solid", fgColor="FDE8E8")}
    ws.append(APPROVED_COLUMNS)
    for c in range(1, len(APPROVED_COLUMNS) + 1):
        ws.cell(1, c).fill = hf
        ws.cell(1, c).font = ht
        ws.cell(1, c).alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(c, "") for c in APPROVED_COLUMNS])
        last = ws.max_row
        f = fills.get("RATE" if r["rule_type"] == "RATE" else ("ALLOW" if r["effect"] == "ALLOW" else "DECLINE"))
        if f:
            for c in range(1, len(APPROVED_COLUMNS) + 1):
                ws.cell(last, c).fill = f
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(APPROVED_COLUMNS))}{ws.max_row}"
    widths = {"reason": 42, "source_text": 48, "geo_label": 22, "canonical_state": 18,
              "model": 18, "make": 16, "catalog_id": 20, "source_rule_id": 18}
    for i, col in enumerate(APPROVED_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 12)
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

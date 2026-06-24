"""
Extractor for the GoDigit large-broker motor grid
("Large Broker Grid Jun'26  Godigit.xlsx").

Emits rows in the same 27-column atomic-catalog schema (one source cell == one
catalog row), insurer = GODIGIT.

GoDigit specifics:
  - commission is CD1 (first-level, a cap) + CD2 (second-level payout). The
    catalog pay_in_pct = **Max CD2** for the policy; CD1 and Avg CD2 are kept
    in source_text so nothing is lost.
  - values are a mix of fractions (0.35), percent strings ("50%\\n"), and the
    markers 'D' (declined -> ELIGIBILITY row) and 'MISP' (payable only via the
    MISP program -> warning, not extracted as a rate).
  - geography is per-product cluster vocabularies; geo_maps() merges the
    'Updated Connect_RTO Mapping' + school/staff-bus RTO mapping sheets.

Sheets extracted: CV Grid (excl. HCV) flat table, HCV Grid, Pvt Car TP Grid,
Pvt Car Comp+SAOD Grid, 2W (1+1 & SATP, 1+5, 5+5, SAOD), School Bus, Staff Bus.
"""
import glob
import re

import openpyxl
from openpyxl.utils import get_column_letter

from catalog_schema import row
import geo_lookup as gn

INSURER = "GODIGIT"


def _file():
    return glob.glob("Large Broker Grid*Godigit.xlsx")[0]


def pnum(v):
    """Cell -> percent number | 'D' | 'MISP' | None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return round(f * 100, 3) if abs(f) <= 1.0 else round(f, 3)
    s = str(v).strip().replace("\n", "")
    u = s.upper()
    if u in ("D", "DECLINE", "DECLINED"):
        return "D"
    if u == "MISP":
        return "MISP"
    m = re.fullmatch(r"(-?\d+(\.\d+)?)\s*%", s)
    if m:
        return round(float(m.group(1)), 3)
    m = re.fullmatch(r"-?\d+(\.\d+)?", s)
    if m:
        f = float(s)
        return round(f * 100, 3) if abs(f) <= 1.0 else round(f, 3)
    return None


def _age(lo, hi):
    out = {}
    if isinstance(lo, (int, float)):
        out["age_min"] = int(lo)
    if isinstance(hi, (int, float)) and hi < 99:
        out["age_max"] = int(hi)
    return out


def _cat(seg):
    s = str(seg).upper()
    if s.startswith("GCV") or "E-LOADER" in s or "E LOADER" in s:
        return "GCV"
    if (s.startswith("PCV") or "E-RICKSHAW" in s or "E RICKSHAW" in s
            or "E-AUTO" in s or "E AUTO" in s or "BUS" in s or "TAXI" in s
            or "WINGER" in s or "MAXI" in s or "SEATER" in s):
        return "PCV"
    if "TRACTOR" in s or "MISC" in s or "BACKHOE" in s or "EXCAVATOR" in s or "JCB" in s:
        return "MISD"
    return ""


_AGE_PATTERNS = [
    (re.compile(r"(\d+)\s*(?:to|-)\s*(\d+)\s*year", re.I), lambda m: (int(m.group(1)), int(m.group(2)))),
    (re.compile(r"(\d+)\s*\+\s*year", re.I), lambda m: (int(m.group(1)), None)),
    (re.compile(r"age\s*(\d+)\s*(?:to|-)\s*(\d+)", re.I), lambda m: (int(m.group(1)), int(m.group(2)))),
    (re.compile(r"age\s*(\d+)\s*\+", re.I), lambda m: (int(m.group(1)), None)),
    (re.compile(r"greater than\s*(\d+)\s*year", re.I), lambda m: (int(m.group(1)) + 1, None)),
    (re.compile(r"age\s*(\d+)\b", re.I), lambda m: (int(m.group(1)), int(m.group(1)))),
    (re.compile(r"\b(\d+)\s*year(?:s)?\b", re.I), lambda m: (int(m.group(1)), int(m.group(1)))),
]


def _seg_age(seg):
    """Pull an age band out of free-text segment labels like
    'GCV4 upto 2.5T 3+ years' or 'PCV3W non-diesel Age 1-2'."""
    for pat, fn in _AGE_PATTERNS:
        m = pat.search(seg)
        if m:
            lo, hi = fn(m)
            return {"age_min": lo, **({} if hi is None else {"age_max": hi})}
    return {}


_AGELINE = re.compile(
    r"(?:Age\s*)?(?:(>=?)\s*(\d+)|(\d+)\s*\+|(\d+)(?:\s*(?:to|-)\s*(\d+))?)"
    r"\s*(?:yr|yrs|year|years|age)?\s*[-:]\s*(\d+(?:\.\d+)?)\s*%", re.I)


# GoDigit CV-grid block titles ('Good Vizag_Vijayawada', 'Bad KA', 'Good UP', …)
# don't fold to a state and drift in spelling from the RTO-map vocabulary, so they
# were unreachable by state. Map the title's state token to a canonical state so
# the resolver can match them by state (in addition to exact cluster typing).
_CL_STATE_TOKENS = [
    ("VIZAG", "ANDHRA PRADESH"), ("VIJAYWADA", "ANDHRA PRADESH"), ("VIJAYAWADA", "ANDHRA PRADESH"),
    ("BANGALORE", "KARNATAKA"), ("KARNATAKA", "KARNATAKA"), ("BAD KA", "KARNATAKA"),
    ("KA GOOD", "KARNATAKA"), ("KA REF", "KARNATAKA"), (" KA ", "KARNATAKA"),
    ("KERALA", "KERALA"), ("LAKSHWADEEP", "KERALA"), ("LAKSHADWEEP", "KERALA"),
    ("BHUBUNESHWAR", "ORISSA"), ("ORISSA", "ORISSA"), ("ODISHA", "ORISSA"),
    ("PATNA", "BIHAR"), ("BIHAR", "BIHAR"),
    ("RANCHI", "JHARKHAND"), ("JHARKHAND", "JHARKHAND"),
    ("LUCKNOW", "UTTAR PRADESH"), ("GOOD UP", "UTTAR PRADESH"), ("BAD UP", "UTTAR PRADESH"),
    ("UP CLUSTER", "UTTAR PRADESH"),
    ("GOOD UK", "UTTARAKHAND"), ("BAD UK", "UTTARAKHAND"),
    ("JAMMU", "JAMMU AND KASHMIR"), ("SRINAGAR", "JAMMU AND KASHMIR"),
    ("HIMACHAL", "HIMACHAL PRADESH"), ("HP REF", "HIMACHAL PRADESH"),
    ("PUNJAB", "PUNJAB"), ("PB REF", "PUNJAB"),
    ("HR REF", "HARYANA"), ("HARYANA", "HARYANA"),
    ("DELHI", "DELHI"), ("NCR", "DELHI"),
    ("MUMBAI", "MAHARASHTRA"), ("PUNE", "MAHARASHTRA"), ("ROM", "MAHARASHTRA"),
    ("MH REF", "MAHARASHTRA"),
    ("GOOD RJ", "RAJASTHAN"), ("BAD RJ", "RAJASTHAN"), ("RJ REF", "RAJASTHAN"),
    ("ASSAM", "ASSAM"), ("ANDAMAN", "ANDAMAN & NICOBAR ISLANDS"),
    ("GOA", "GOA"), ("GUJARAT", "GUJARAT"),
]


def _cluster_state(label):
    """Single canonical state for a CV-grid block-title cluster, or '' if it
    spans several states / can't be hinted (kept blank rather than guessed)."""
    up = " " + str(label or "").upper() + " "
    if "MP" in up and "CG" in up:
        return ""        # 'MP & CG' spans two states
    if "AP" in up and "TS" in up:
        return ""        # 'Bad AP&TS' spans two states
    for tok, st in _CL_STATE_TOKENS:
        if tok in up:
            return st
    if " MP " in up:
        return "MADHYA PRADESH"
    if " CG " in up:
        return "CHATTISGARH"
    return gn.to_state(label) or ""


def _merged_value(ws, r, c):
    """Cell value, resolving a blank continuation row of a MERGED cell to the
    range's anchor value (so merged RTO-cluster cells don't drop rows)."""
    v = ws.cell(r, c).value
    if v is not None and str(v).strip():
        return v
    for m in ws.merged_cells.ranges:
        if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
            return ws.cell(m.min_row, m.min_col).value
    return v


def _slash_rates(raw):
    """'52%/44.5%' or '50%/32.5%/17.5%' -> [52.0, 44.5] / [50,32.5,17.5]. [] if not that shape."""
    s = str(raw).strip()
    if "/" not in s or "AGE" in s.upper():
        return []
    out = []
    for part in s.split("/"):
        m = re.search(r"(\d+(?:\.\d+)?)\s*%?", part)
        if m:
            out.append(float(m.group(1)))
    return out if len(out) >= 2 else []


def age_split_cell(raw):
    """'Age 0 to 5-7.5%\\nAge>=6: 20%' -> [(0,5,7.5), (6,None,20.0)]. [] if no match."""
    out = []
    for m in _AGELINE.finditer(str(raw)):
        pctv = float(m.group(6))
        if m.group(1):           # Age>=N
            out.append((int(m.group(2)), None, pctv))
        elif m.group(3):         # Age N+
            out.append((int(m.group(3)), None, pctv))
        else:                    # Age N [to M]
            lo = int(m.group(4))
            hi = int(m.group(5)) if m.group(5) else lo
            out.append((lo, hi, pctv))
    return out


def extract_rows(src=None):
    wb = openpyxl.load_workbook(src or _file(), data_only=True)
    rate_rows, elig_rows, warnings = [], [], []
    seq = {"CV": 0, "HCV": 0, "PCTP": 0, "PC": 0, "TW": 0, "BUS": 0, "ELIG": 0}
    n_misp = 0

    def nid(k):
        seq[k] += 1
        return f"DIGIT-{k}-{seq[k]:04d}"

    n_neg = []

    def emit(k, **kw):
        kw.setdefault("insurer", INSURER)
        kw.setdefault("rule_type", "RATE")
        kw.setdefault("effect", "RATE")
        kw.setdefault("applies_on", "NET")
        kw.setdefault("geo_kind", "RTO_CLUSTER")
        kw.setdefault("confidence", 1.0)
        kw.setdefault("review_status", "PENDING")
        # a negative Max CD2 (CD2-reduction column leaked as a pay-in) is never a
        # valid commission — clamp to 0 and flag rather than serving a negative %.
        if isinstance(kw.get("pay_in_pct"), (int, float)) and kw["pay_in_pct"] < 0:
            n_neg.append((kw.get("source_cell", ""), kw["pay_in_pct"]))
            kw["pay_in_pct"] = 0.0
            kw["review_status"] = "NEEDS_REVIEW"
        rate_rows.append(row(catalog_id=nid(k), **kw))

    def decline(k, sheet, cell, cat, sub, mk, pol, geo_label, why):
        elig_rows.append(row(
            catalog_id=nid("ELIG"), source_rule_id=f"{sheet}!{cell}",
            insurer=INSURER, rule_type="ELIGIBILITY", effect="DECLINE",
            category=cat, sub_segment=sub, make=mk, policy_type=pol,
            geo_kind="RTO_CLUSTER", geo_label=geo_label,
            reason=why, source_sheet=sheet, source_cell=cell,
            source_text=why, confidence=1.0, review_status="PENDING"))

    # ------------------------------------------------ CV Grid: per-cluster left blocks
    # Block anatomy: a title row in col B (the cluster), a header row
    # (Segment | Make | [Carrier Type | Addon...] | CD1 | Avg CD2 | Max CD2
    # [| SATP Avg CD2 | Max CD2]), then data rows. Layout varies per block, so the
    # column map is rebuilt at every header row. Unparseable cells (nested age
    # sub-tables, slab text) are surfaced as warnings, never guessed.
    SH = "CV Grid (excl. HCV)"
    ws = wb[SH]
    cluster, cmap, n_skip = "", None, 0
    last_seg, last_mk = "", ""
    for r in range(2, ws.max_row + 1):
        cells = [(c, str(ws.cell(r, c).value).strip()) for c in range(2, 12)
                 if ws.cell(r, c).value is not None]
        tokens = {v.upper(): c for c, v in cells}
        # header row: MAKE alone is enough (some HCV-style sub-tables have NO
        # 'Segment' column — make-only). Build a TOKEN-positioned column map so
        # leading-'Cluster' layouts (Cluster|Segment|Make|...) are read correctly.
        if "MAKE" in tokens and ("CD1" in tokens or "MAX CD2" in tokens):
            avgs = sorted(c for c, v in cells if v.upper() == "AVG CD2")
            maxs = sorted(c for c, v in cells if v.upper() == "MAX CD2")
            cmap = {"seg": tokens.get("SEGMENT"), "make": tokens.get("MAKE"),
                    "rowcluster": next((c for t, c in tokens.items() if t == "CLUSTER"), None),
                    "cd1": tokens.get("CD1"),
                    "carrier": next((c for t, c in tokens.items() if "CARRIER" in t), None),
                    "addon": next((c for t, c in tokens.items() if "ADDON" in t or "ADD ON" in t), None),
                    "comp": (avgs[0] if avgs else None, maxs[0] if maxs else None),
                    "satp": (avgs[1] if len(avgs) > 1 else None, maxs[1] if len(maxs) > 1 else None)}
            # block-title cluster: nearest non-empty B above the header that is not
            # itself a header/footnote
            for rr in range(r - 1, max(1, r - 4), -1):
                t = ws.cell(rr, 2).value
                if t and str(t).strip() and not re.search(r"Segment|Make|No need to upload", str(t)):
                    cluster = re.sub(r"\s+(CV Grid|RTO Cluster|Grid)$", "", str(t).strip(), flags=re.I)
                    break
            last_seg = last_mk = ""
            continue
        if not cmap:
            continue
        seg_col, mk_col = cmap["seg"], cmap["make"]
        seg = (str(ws.cell(r, seg_col).value).strip() if seg_col and ws.cell(r, seg_col).value
               else (last_seg if seg_col else cluster))  # make-only block -> title is the segment
        mk = str(ws.cell(r, mk_col).value or "").strip() if mk_col else ""
        mk = mk or last_mk
        if not seg:
            continue
        cat = _cat(seg)
        if not cat:
            continue
        last_seg, last_mk = seg, mk
        # per-row geo: a 'Cluster' column overrides the block title
        rowgeo = (str(ws.cell(r, cmap["rowcluster"]).value or "").strip()
                  if cmap["rowcluster"] else "") or cluster
        carrier = str(ws.cell(r, cmap["carrier"]).value or "").strip() if cmap["carrier"] else ""
        addon = str(ws.cell(r, cmap["addon"]).value or "").strip() if cmap["addon"] else ""
        cd1 = ws.cell(r, cmap["cd1"]).value if cmap["cd1"] else None
        sub = re.sub(r"\s+", " ", seg)
        if carrier and carrier.upper() != "ALL":
            sub += f" [{carrier}]"
        if addon:
            sub += f" [{addon[:30]}]"
        agekw = _seg_age(seg)
        for (cavg, cmax), pol in ((cmap["comp"], "COMP"), (cmap["satp"], "SATP")):
            if not cmax:
                continue
            raw = ws.cell(r, cmax).value
            val = pnum(raw)
            cell = f"{get_column_letter(cmax)}{r}"
            if raw is not None and val is None:
                bands = age_split_cell(raw)  # in-cell age-banded values
                for lo, hi, p in bands:
                    emit("CV", source_rule_id=f"{SH}!{cell}", pay_in_pct=p,
                         category=cat, sub_segment=sub,
                         make=(mk if mk.upper() != "ALL" else ""), policy_type=pol,
                         age_min=lo, **({} if hi is None else {"age_max": hi}),
                         geo_label=rowgeo, canonical_state=_cluster_state(rowgeo), source_sheet=SH, source_cell=cell,
                         source_text=f"{seg} {pol}; in-cell band: {str(raw)[:60]}")
                if bands:
                    continue
                slashed = _slash_rates(raw)  # 'NCB/non-NCB' multi-rate cell
                if slashed:
                    emit("CV", source_rule_id=f"{SH}!{cell}", pay_in_pct=min(slashed),
                         category=cat, sub_segment=sub, make=(mk if mk.upper() != "ALL" else ""),
                         policy_type=pol, **agekw, geo_label=rowgeo, canonical_state=_cluster_state(rowgeo), source_sheet=SH,
                         source_cell=cell,
                         source_text=f"{seg} {pol}; multi-rate {str(raw)[:40]} (kept min, conservative)")
                    warnings.append({"scope": SH, "cell": cell,
                                     "issue": f"multi-rate cell {str(raw)[:40]!r} ({rowgeo}/{seg}) "
                                              "— kept the conservative (min) value; confirm tiering"})
                    continue
                n_skip += 1
                warnings.append({"scope": SH, "cell": cell,
                                 "issue": f"unparseable rate cell {str(raw)[:50]!r} ({rowgeo} / {seg})"})
                continue
            if val is None:
                continue
            if val == "D":
                decline("CV", SH, cell, cat, sub, mk if mk.upper() != "ALL" else "",
                        pol, rowgeo, f"Declined: {seg} {pol} @ {rowgeo} (grid 'D')")
                continue
            if val == "MISP":
                n_misp += 1
                continue
            avg = pnum(ws.cell(r, cavg).value) if cavg else None
            emit("CV", source_rule_id=f"{SH}!{cell}", pay_in_pct=val,
                 category=cat, sub_segment=sub, make=(mk if mk.upper() != "ALL" else ""),
                 policy_type=pol, **agekw, geo_label=rowgeo, canonical_state=_cluster_state(rowgeo),
                 source_sheet=SH, source_cell=cell,
                 source_text=f"{seg} {pol}; CD1 {str(cd1)[:30]}; Avg CD2 {avg}")
    if n_skip:
        warnings.append({"scope": SH, "cell": "various",
                         "issue": f"{n_skip} rate cells were text (nested age sub-tables / "
                                  "slab notes) and need manual review"})

    # ------------------------------------------------ HCV Grid
    SH = "HCV Grid"
    ws = wb[SH]
    BODY = [("NON-DUMPER/TIPPER", 7), ("DUMPER/TIPPER", 12), ("OIL TANKER", 17), ("GAS TANKER", 22)]
    for r in range(4, ws.max_row + 1):
        cluster, seg, mk = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        # GODIGIT-7: banner rows (cluster present, seg blank) carry make-block /
        # CD1-cap eligibility notes — surface them as warnings instead of dropping.
        if cluster and not seg:
            note = " ".join(str(ws.cell(r, c).value) for c in range(2, 10)
                            if ws.cell(r, c).value is not None)
            if re.search(r"block|decline|cap|only", note, re.I):
                warnings.append({"scope": SH, "cell": f"row {r}",
                                 "issue": f"HCV banner note not modelled as a rule: {note[:120]}"})
            continue
        if not (cluster and seg) or str(seg).strip() == "Segment":
            continue
        cluster, seg = str(cluster).strip(), str(seg).strip()
        agekw = _age(ws.cell(r, 5).value, ws.cell(r, 6).value)
        for body, c0 in BODY:
            cd1 = ws.cell(r, c0).value
            for off, pol in ((2, "COMP"), (4, "SATP")):  # Max CD2 cols: c0+2 comp, c0+4 satp
                val = pnum(ws.cell(r, c0 + off).value)
                if val in (None, "D", "MISP"):
                    if val == "D":
                        decline("HCV", SH, f"{get_column_letter(c0+off)}{r}", "GCV",
                                f"{seg} [{body}]", "", pol, cluster, f"Declined: {seg} {body} {pol} @ {cluster}")
                    continue
                emit("HCV", source_rule_id=f"{SH}!{get_column_letter(c0+off)}{r}",
                     pay_in_pct=val, category="GCV", sub_segment=f"{seg} [{body}]",
                     make=(str(mk).strip() if mk and str(mk).strip().upper() != "ALL" else ""),
                     policy_type=pol, **agekw, geo_label=cluster,
                     source_sheet=SH, source_cell=f"{get_column_letter(c0+off)}{r}",
                     source_text=f"{seg} {body} {pol}; CD1 {str(cd1)[:40]}")

    # ------------------------------------------------ Pvt Car TP
    SH = "Pvt Car TP Grid"
    ws = wb[SH]
    fuel_re = re.compile(r"^(Petrol|Diesel|CNG|EV|Electric)\s*(?:([<>])\s*(\d+)|(\d+)\s*-\s*(\d+))?\s*$", re.I)
    for r in range(3, ws.max_row + 1):
        cluster, seg, age = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        val = pnum(ws.cell(r, 5).value)
        if not (cluster and seg) or val is None:
            continue
        m = fuel_re.match(str(seg).strip())
        if not m:
            warnings.append({"scope": SH, "cell": f"C{r}", "issue": f"unknown PC TP segment {seg!r}"})
            continue
        fuel = m.group(1).upper()
        lo = hi = None
        if m.group(2) == "<":
            lo, hi = 0, int(m.group(3)) - 1
        elif m.group(2) == ">":
            lo = int(m.group(3)) + 1
        elif m.group(4):
            lo, hi = int(m.group(4)), int(m.group(5))
        sub = f"PC TP [{fuel}]"
        a = str(age or "All").strip()
        agekw = {} if a == "All" else ({"age_min": 0, "age_max": 9} if a == "<10" else {"age_min": 10})
        if val == "D":
            decline("PCTP", SH, f"E{r}", "PVT_CAR", sub, "", "SATP", str(cluster).strip(),
                    f"Declined: {sub} age {a} @ {cluster}")
            continue
        emit("PCTP", source_rule_id=f"{SH}!E{r}", pay_in_pct=val, category="PVT_CAR",
             sub_segment=sub, policy_type="SATP",
             cc_min=("" if lo is None else lo), cc_max=("" if hi is None else hi), **agekw,
             geo_label=str(cluster).strip(), source_sheet=SH, source_cell=f"E{r}",
             source_text=f"{seg} age {a}")

    # ------------------------------------------------ Pvt Car Comp+SAOD
    SH = "Pvt Car Comp+SAOD Grid"
    ws = wb[SH]
    nb_makes = []
    for c in range(3, 10):
        h = ws.cell(5, c).value
        if h:
            h = str(h).split("\n")[0].replace("_NB", "").strip()
            nb_makes.append((c, "HEV MAKES" if h.upper().startswith("HEV") else
                             ("OTHER MAKES" if h.upper().startswith("OTHERS") else h)))
    for r in range(6, ws.max_row + 1):
        cluster = ws.cell(r, 2).value
        if cluster and str(cluster).strip():
            cl = str(cluster).strip()
            for c, mk in nb_makes:
                val = pnum(ws.cell(r, c).value)
                if val is None or val in ("D", "MISP"):
                    continue
                emit("PC", source_rule_id=f"{SH}!{get_column_letter(c)}{r}", pay_in_pct=val,
                     category="PVT_CAR", sub_segment="PC NEW BUSINESS (1+3)", make=mk,
                     policy_type="COMP-NB", geo_label=cl, source_sheet=SH,
                     source_cell=f"{get_column_letter(c)}{r}", source_text=f"NB 1+3 90:10 {mk}")
            v33 = pnum(ws.cell(r, 10).value)  # J: 3+3 payable over 3 yrs
            if v33 not in (None, "D", "MISP"):
                emit("PC", source_rule_id=f"{SH}!J{r}", pay_in_pct=v33,
                     category="PVT_CAR", sub_segment="PC NEW BUSINESS (3+3)", make="",
                     policy_type="COMP-NB", geo_label=cl, source_sheet=SH,
                     source_cell=f"J{r}", source_text="NB 3+3 payable over 3 years")
    RENEW = [(15, "SAOD", "PC RENEWAL [NON-NCB]", ""), (16, "SAOD", "PC RENEWAL [NCB]", ""),
             (17, "COMP", "PC RENEWAL [NON-NCB]", ""), (18, "COMP", "PC RENEWAL [NCB]", ""),
             (19, "COMP", "PC RENEWAL HEV", "HEV MAKES")]
    for r in range(5, ws.max_row + 1):
        cluster = ws.cell(r, 14).value
        if not (cluster and str(cluster).strip()) or str(cluster).strip() == "Cluster":
            continue
        cl = str(cluster).strip()
        for c, pol, sub, mk in RENEW:
            val = pnum(ws.cell(r, c).value)
            if val is None or val in ("D", "MISP"):
                continue
            emit("PC", source_rule_id=f"{SH}!{get_column_letter(c)}{r}", pay_in_pct=val,
                 category="PVT_CAR", sub_segment=sub, make=mk, policy_type=pol,
                 geo_label=cl, source_sheet=SH, source_cell=f"{get_column_letter(c)}{r}",
                 source_text=f"Renewal {pol} {sub}")

    # ------------------------------------------------ 2W grids
    SH = "2W Grid 1+1 & SATP"
    ws = wb[SH]
    for r in range(4, ws.max_row + 1):
        cluster, seg = ws.cell(r, 2).value, ws.cell(r, 3).value
        if not (cluster and seg):
            continue
        cl, sub = str(cluster).strip(), str(seg).strip()
        cd1 = pnum(ws.cell(r, 4).value)
        for c, pol in ((5, "COMP (1+1)"), (7, "SATP")):
            val = pnum(ws.cell(r, c).value)
            if val is None:
                continue
            if val == "D":
                decline("TW", SH, f"{get_column_letter(c)}{r}", "TW", sub, "", pol, cl,
                        f"Declined: {sub} {pol} @ {cl}")
                continue
            if val == "MISP":
                n_misp += 1
                continue
            emit("TW", source_rule_id=f"{SH}!{get_column_letter(c)}{r}", pay_in_pct=val,
                 category="TW", sub_segment=sub, policy_type=pol, geo_label=cl,
                 source_sheet=SH, source_cell=f"{get_column_letter(c)}{r}",
                 source_text=f"{sub} {pol}; CD1 {cd1}")

    for SH, cols in (("2W Grid 1+5", {"cl": 2, "mk": 3, "seg": 4, "cd1": 5, "max": 7, "pol": "COMP (1+5)", "r0": 3}),
                     ("2W Grid 5+5", {"cl": 3, "mk": 4, "seg": 5, "cd1": 6, "max": 8, "pol": "COMP (5+5)", "r0": 3})):
        ws = wb[SH]
        for r in range(cols["r0"], ws.max_row + 1):
            cluster, mk, seg = (ws.cell(r, cols["cl"]).value, ws.cell(r, cols["mk"]).value,
                                ws.cell(r, cols["seg"]).value)
            if not (cluster and seg):
                continue
            cl, sub = str(cluster).strip(), str(seg).strip()
            mk = str(mk or "").strip()
            val = pnum(ws.cell(r, cols["max"]).value)
            if val is None:
                continue
            if val == "D":
                decline("TW", SH, f"{get_column_letter(cols['max'])}{r}", "TW", sub, mk,
                        cols["pol"], cl, f"Declined: {mk} {sub} {cols['pol']} @ {cl}")
                continue
            if val == "MISP":
                n_misp += 1
                continue
            emit("TW", source_rule_id=f"{SH}!{get_column_letter(cols['max'])}{r}",
                 pay_in_pct=val, category="TW", sub_segment=sub,
                 make=(mk if mk.upper() != "ALL" else ""), policy_type=cols["pol"],
                 geo_label=cl, source_sheet=SH,
                 source_cell=f"{get_column_letter(cols['max'])}{r}",
                 source_text=f"{mk} {sub} {cols['pol']}; CD1 {pnum(ws.cell(r, cols['cd1']).value)}")

    SH = "2W_SAOD"
    ws = wb[SH]
    for r in range(3, ws.max_row + 1):
        cluster, seg = ws.cell(r, 2).value, ws.cell(r, 3).value
        if not (cluster and seg):
            continue
        cl, sub = str(cluster).strip(), str(seg).strip()
        for c, yr in ((7, 1), (8, 2), (9, 3), (10, 4)):
            val = pnum(ws.cell(r, c).value)
            if val in (None, "D", "MISP"):
                continue
            emit("TW", source_rule_id=f"{SH}!{get_column_letter(c)}{r}", pay_in_pct=val,
                 category="TW", sub_segment=sub, policy_type="SAOD",
                 age_min=yr, age_max=yr, geo_label=cl, source_sheet=SH,
                 source_cell=f"{get_column_letter(c)}{r}",
                 source_text=f"{sub} SAOD year {yr}")

    # ------------------------------------------------ School Bus
    # Layout: State_1 (col B, MERGED — forward-fill) | RTO Cluster (col C) |
    # Seating (col D) | rates E..H. Continuation rows leave col B blank, so the
    # cluster sits in col C with B empty; the old code skipped them and lost most
    # rows (Rest of AP, TS Open (Others), Gujarat Others, KA Open, …).
    SH = "School Bus - Agency"
    ws = wb[SH]
    # sub_segment naming is the original stable scheme ("SCHOOL BUS 8+ [scope]");
    # the seating band (col D) is kept in source_text, NOT the label, so the
    # filter values stay constant month to month.
    SB = [(5, "SCHOOL BUS 8+ [IN SCHOOL NAME]", "COMP (VOL 1-3L)"),
          (6, "SCHOOL BUS 8+ [IN SCHOOL NAME]", "COMP (VOL >3L)"),
          (7, "SCHOOL BUS 8+ [ON CONTRACT - TRANSPORTER]", "COMP"),
          (8, "SCHOOL BUS 8+ [ON CONTRACT - INDIVIDUAL]", "COMP")]
    state = None
    for r in range(6, ws.max_row + 1):
        b = ws.cell(r, 2).value
        cl = _merged_value(ws, r, 3)  # GODIGIT-6: resolve merged RTO-cluster cells
        if b and str(b).strip():
            state = str(b).strip()
        cluster = str(cl).strip() if cl and str(cl).strip() else None
        if not cluster:
            continue
        if "Term" in cluster or "Condition" in cluster:  # footer block
            break
        seat = str(ws.cell(r, 4).value or "").strip()
        cstate = gn.to_state(cluster) or gn.to_state(state) or ""
        for c, sub, pol in SB:
            raw = ws.cell(r, c).value
            val = pnum(raw)
            cell = f"{get_column_letter(c)}{r}"
            if val == "D" or (isinstance(raw, str) and raw.strip().upper() == "D"):
                decline("BUS", SH, cell, "PCV", sub, "", pol, cluster,
                        f"Declined: {sub} @ {state}/{cluster}")
                continue
            if val in (None, "MISP"):
                continue
            emit("BUS", source_rule_id=f"{SH}!{cell}", pay_in_pct=val,
                 category="PCV", sub_segment=sub, policy_type=pol, geo_label=cluster,
                 canonical_state=cstate, source_sheet=SH, source_cell=cell,
                 source_text=f"{sub} {pol} @ {state}/{cluster} (seat {seat})")

    # ------------------------------------------------ Staff Bus (text cells)
    # Same merged State_1 (col B); cluster in col C; values are "CD1 95% / CD2 X%"
    # text or "Decline". Continuation rows (Vijaywada, Pune, ROM, …) were dropped.
    SH = "Staff Bus Grid (2)"
    ws = wb[SH]
    STB = [(4, "STAFF BUS [CORPORATE SELF-USAGE]"), (5, "STAFF BUS [CONTRACT - TRANSPORTER]"),
           (6, "STAFF BUS [CONTRACT - INDIVIDUAL]")]
    cd_re = re.compile(r"CD2\s*([\d.]+)\s*%", re.I)
    state = None
    for r in range(5, ws.max_row + 1):
        b = ws.cell(r, 2).value
        cl = _merged_value(ws, r, 3)
        if b and str(b).strip():
            state = str(b).strip()
            if state.lower().startswith("note"):
                break
        cluster = str(cl).strip() if cl and str(cl).strip() else None
        if not cluster:
            continue
        cstate = gn.to_state(cluster) or gn.to_state(state) or ""
        for c, sub in STB:
            raw = ws.cell(r, c).value
            if raw is None:
                continue
            txt = str(raw).strip()
            cell = f"{get_column_letter(c)}{r}"
            m = cd_re.search(txt)
            if m:
                emit("BUS", source_rule_id=f"{SH}!{cell}",
                     pay_in_pct=float(m.group(1)), category="PCV", sub_segment=sub,
                     policy_type="COMP", geo_label=cluster, canonical_state=cstate,
                     source_sheet=SH, source_cell=cell, source_text=f"{txt[:60]} @ {state}/{cluster}")
            elif txt.upper().startswith("DECLINE"):
                decline("BUS", SH, cell, "PCV", sub, "", "COMP",
                        cluster, f"Declined: {sub} @ {state}/{cluster}")

    if n_misp:
        warnings.append({"scope": "workbook", "cell": "various",
                         "issue": f"{n_misp} cells are 'MISP' (payable only via MISP program, "
                                  "no broker commission) — not extracted as rates. A quote for "
                                  "these (make, segment, cluster) returns NO_RATE; confirm whether "
                                  "they should be modelled as DECLINE for this channel."})
    if n_neg:
        warnings.append({"scope": "workbook", "cell": ", ".join(c for c, _ in n_neg[:8]),
                         "issue": f"{len(n_neg)} rate cell(s) had a NEGATIVE Max CD2 (a CD2-reduction "
                                  "value leaked into the pay-in column) — clamped to 0 and flagged "
                                  "review_status=NEEDS_REVIEW; confirm the correct payout."})
    return rate_rows, elig_rows, warnings


def geo_maps(src=None):
    """Merge all GoDigit RTO->cluster mapping sheets into rto2clusters / cluster2states."""
    wb = openpyxl.load_workbook(src or _file(), data_only=True)
    rto2clusters, cluster2states = {}, {}
    from collections import defaultdict
    pre = defaultdict(set)
    for code, st in gn.CODE2STATE.items():
        pre[code[:2]].add(st)
    prefix2state = {p: next(iter(s)) for p, s in pre.items() if len(s) == 1}

    def add(code, label, state_hint=None):
        if not code or not label:
            return
        code = str(code).strip().upper()
        label = re.sub(r"\s+", " ", str(label).strip())
        # GEO-06: strip CD-band/discount suffixes and reject non-geographic labels
        # (CD-band tags, decline tags, footnotes) so they don't pollute the vocab.
        label = re.sub(r"_+LOWCD2.*$", "", label, flags=re.I).strip()
        if (not label or len(label) > 40 or "LOWCD2" in label.upper()
                or re.search(r"(block|decline|declined|upload|cap at|only for|note|all_india)",
                             label, re.I)):
            return
        if not re.fullmatch(r"[A-Z]{2}\d{1,3}[A-Z]?", code):
            return
        rto2clusters.setdefault(code, set()).add(label)
        # prefer the mapping sheet's own State column, then the RTO master, then prefix
        st = (gn.to_state(state_hint) if state_hint else None) \
            or gn.CODE2STATE.get(code) or prefix2state.get(code[:2])
        if st:
            cluster2states.setdefault(label, set()).add(st)

    ws = wb["Updated Connect_RTO Mapping"]
    for r in range(3, ws.max_row + 1):
        for code_c, label_cs in ((2, range(3, 11)), (13, [14]), (16, [17, 18]),
                                 (20, [21]), (23, [24]), (26, [27])):
            code = ws.cell(r, code_c).value
            for lc in label_cs:
                add(code, ws.cell(r, lc).value)
    # bus mapping sheets: code | State | State_1 | granular-cluster. The granular
    # cluster (last col) is what the bus grids key on; State_1 is a coarser group.
    # Both are registered, with the State column used to fold to a canonical state.
    for sheet, code_c, state_c, label_cs, r0 in (
            ("RTO Mapping-School bus", 2, 3, [5, 4], 3),
            ("RTO Mapping -Staff bus", 1, 2, [4, 3], 2)):
        ws = wb[sheet]
        for r in range(r0, ws.max_row + 1):
            code = ws.cell(r, code_c).value
            state_hint = ws.cell(r, state_c).value
            for lc in label_cs:
                add(code, ws.cell(r, lc).value, state_hint)
    return ({c: sorted(s) for c, s in rto2clusters.items()},
            {c: sorted(s) for c, s in cluster2states.items()})


if __name__ == "__main__":
    rates, eligs, warns = extract_rows()
    from collections import Counter
    print(f"GODIGIT: {len(rates)} RATE, {len(eligs)} ELIG, {len(warns)} warnings")
    print("  by category:", dict(Counter(r["category"] for r in rates)))
    print("  by sheet   :", dict(Counter(r["source_sheet"] for r in rates)))
    print("  by policy  :", dict(Counter(r["policy_type"] for r in rates)))
    r2c, c2s = geo_maps()
    print(f"  geo: {len(r2c)} rto codes, {len(c2s)} cluster labels")
    for w in warns:
        print("  WARN", str(w)[:120])

"""
Extractor for 4-Wheeler (Private Car) and Commercial Vehicle (GCV / PCV / MISD)
pay-in rules from the SBI General grid.

Emits rows in the SAME 27-column atomic-catalog schema as build_rule_catalog.row()
so they merge straight into rules_catalog.* and api/_data/catalog.json. One source
grid cell == one catalog row (faithful, no invented policy splits). New physical
dimensions that the grid keys on (GVW weight band, make-tier, fuel, nil-dep, age
band, carrying capacity) are carried in the existing columns:

  sub_segment      segment + weight/seating label  (e.g. "GCV 12-20T", "PCV TAXI")
  make             make-tier when the grid splits on it ("TATA & ASHOK LEYLAND",
                   "OTHER MAKES", "MAHINDRA (ALL VARIANTS)") or actual make (highend)
  policy_type      COMP | SATP | COMP & SATP | COMP-NIL DEP | COMP-NON NIL DEP | SAOD
                   | COMP (PACKAGE)
  cc_min/cc_max    taxi / SATP cc bands
  age_min/age_max  New=(0,0), Upto-5=(0,5), 1-5=(1,5), Above-5=(5,None), Non-New=(1,None)
  geo_kind         RTO_CLUSTER | STATE_OR_CITY ; geo_label = source cluster/state label

Source sheets:
  'GCV & Pvt Car Payout Condition'  GCV weight matrix + Pvt Car COMP/SAOD (by cluster)
  'PCV, MISD & TW'                  MISD agri-tractor, PCV 3W / Taxi / School Bus (by state)
  'PV SATP Apr-26'                  Pvt Car SATP (fuel x cc, by cluster)
  'PV SAOD- HIGHEND Enabler'        Pvt Car high-end SAOD/Package (make/model cluster)
  'Pvt Car Declined Make & Models'  Pvt Car eligibility DECLINE list
"""
import glob
import re

import openpyxl
from openpyxl.utils import get_column_letter

from build_rule_catalog import row, INSURER  # noqa: F401  (row sets insurer + COLUMNS)
from geo_normalize import to_state
from rto_clusters import cluster_state

FILE = glob.glob("Provincial sbi Grid*.xlsx")[0]
SRC_GCV = "GCV & Pvt Car Payout Condition"
SRC_PCV = "PCV, MISD & TW"
SRC_SATP = "PV SATP Apr-26"
SRC_HE = "PV SAOD- HIGHEND Enabler"
SRC_DECL = "Pvt Car Declined Make & Models"

# age band shorthand -> (min, max)
NEW = (0, 0)
UPTO5 = (0, 5)
AGE15 = (1, 5)
ABOVE5 = (5, None)
NONNEW = (1, None)


def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", s):
        return float(s)
    return None


def _cell(ws, r, c):
    return ws.cell(row=r, column=c).value


# --------------------------------------------------------------------- specs
# Each spec: (col, category, sub_segment, make_tier, policy_type, age, cc_min, cc_max)
# age is an (min,max) tuple or None.
GCV_SPECS = [
    (5,  "GCV", "GCV 4W <=2.5T",            "ALL MAKES (<=2.0T) & TATA (2.0-2.5T)", "COMP",        NEW,    None, None),
    (6,  "GCV", "GCV 4W <=2.5T",            "ALL MAKES (<=2.0T) & TATA (2.0-2.5T)", "COMP & SATP", UPTO5,  None, None),
    (7,  "GCV", "GCV 4W <=2.5T",            "ALL MAKES (<=2.0T) & TATA (2.0-2.5T)", "COMP & SATP", ABOVE5, None, None),
    (8,  "GCV", "GCV 4W 2.0-2.5T",          "OTHER THAN TATA",                      "COMP",        NEW,    None, None),
    (9,  "GCV", "GCV 4W 2.0-2.5T",          "OTHER THAN TATA",                      "COMP & SATP", UPTO5,  None, None),
    (10, "GCV", "GCV 4W 2.0-2.5T",          "OTHER THAN TATA",                      "COMP & SATP", ABOVE5, None, None),
    (11, "GCV", "GCV 3W",                   "",                                     "COMP",        NEW,    None, None),
    (12, "GCV", "GCV 3W",                   "",                                     "COMP & SATP", UPTO5,  None, None),
    (13, "GCV", "GCV 3W",                   "",                                     "COMP & SATP", ABOVE5, None, None),
    (14, "GCV", "GCV 2.5-3.5T",             "MAHINDRA (ALL VARIANTS)",              "COMP & SATP", UPTO5,  None, None),
    (15, "GCV", "GCV 2.5-3.5T",             "MAHINDRA (ALL VARIANTS)",              "COMP & SATP", ABOVE5, None, None),
    (16, "GCV", "GCV 2.5-3.5T",             "TATA & ASHOK LEYLAND",                 "COMP & SATP", UPTO5,  None, None),
    (17, "GCV", "GCV 2.5-3.5T",             "TATA & ASHOK LEYLAND",                 "COMP & SATP", ABOVE5, None, None),
    (18, "GCV", "GCV TRACTOR",              "",                                     "COMP & SATP", None,   None, None),
    (19, "GCV", "GCV 3.5-5.0T",             "ALL MAKE",                             "COMP & SATP", None,   None, None),
    (20, "GCV", "GCV 5.0-7.5T",             "ALL MAKE",                             "COMP & SATP", None,   None, None),
    (21, "GCV", "GCV 7.5-12T",              "ALL MAKE",                             "COMP & SATP", None,   None, None),
    (22, "GCV", "GCV 12-20T",               "OTHER MAKES",                          "COMP & SATP", NEW,    None, None),
    (23, "GCV", "GCV 12-20T",               "OTHER MAKES",                          "COMP & SATP", AGE15,  None, None),
    (24, "GCV", "GCV 12-20T",               "OTHER MAKES",                          "COMP & SATP", ABOVE5, None, None),
    (25, "GCV", "GCV 12-20T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", NEW,    None, None),
    (26, "GCV", "GCV 12-20T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", AGE15,  None, None),
    (27, "GCV", "GCV 12-20T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", ABOVE5, None, None),
    (28, "GCV", "GCV 20-40T",               "OTHER MAKES",                          "COMP & SATP", NEW,    None, None),
    (29, "GCV", "GCV 20-40T",               "OTHER MAKES",                          "COMP & SATP", AGE15,  None, None),
    (30, "GCV", "GCV 20-40T",               "OTHER MAKES",                          "COMP & SATP", ABOVE5, None, None),
    (31, "GCV", "GCV 20-40T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", NEW,    None, None),
    (32, "GCV", "GCV 20-40T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", AGE15,  None, None),
    (33, "GCV", "GCV 20-40T",               "TATA & ASHOK LEYLAND",                 "COMP & SATP", ABOVE5, None, None),
    (34, "GCV", "GCV >40T",                 "OTHER MAKES",                          "COMP & SATP", UPTO5,  None, None),
    (35, "GCV", "GCV >40T",                 "OTHER MAKES",                          "COMP & SATP", ABOVE5, None, None),
    (36, "GCV", "GCV >40T",                 "TATA & ASHOK LEYLAND",                 "COMP & SATP", UPTO5,  None, None),
    (37, "GCV", "GCV >40T",                 "TATA & ASHOK LEYLAND",                 "COMP & SATP", ABOVE5, None, None),
    (38, "PVT_CAR", "PVT CAR",              "",                                     "COMP",        None,   None, None),
    (39, "PVT_CAR", "PVT CAR",              "",                                     "SAOD",        None,   None, None),
]

PCV_SPECS = [
    (4,  "MISD", "AGRI TRACTOR & HARVESTER", "", "COMP",                  NEW,    None, None, None),
    (5,  "MISD", "AGRI TRACTOR & HARVESTER", "", "COMP & SATP",           NONNEW, None, None, None),
    (6,  "PCV",  "PCV 3W (3+1)",             "", "COMP",                  NEW,    None, None, "NON DIESEL"),
    (7,  "PCV",  "PCV 3W (3+1)",             "", "SATP",                  NEW,    None, None, "NON DIESEL"),
    (8,  "PCV",  "PCV 3W (3+1)",             "", "COMP",                  NONNEW, None, None, "NON DIESEL"),
    (9,  "PCV",  "PCV 3W (3+1)",             "", "SATP",                  NONNEW, None, None, "NON DIESEL"),
    (10, "PCV",  "PCV 3W (3+1)",             "", "COMP",                  NEW,    None, None, "DIESEL"),
    (11, "PCV",  "PCV 3W (3+1)",             "", "SATP",                  NEW,    None, None, "DIESEL"),
    (12, "PCV",  "PCV 3W (3+1)",             "", "COMP",                  NONNEW, None, None, "DIESEL"),
    (13, "PCV",  "PCV 3W (3+1)",             "", "SATP",                  NONNEW, None, None, "DIESEL"),
    (14, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NIL DEP",          None,   0,    999,  None),
    (15, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NON NIL DEP",      None,   0,    999,  None),
    (16, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "SATP",                  None,   0,    999,  None),
    (17, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NIL DEP",          None,   1000, 1499, None),
    (18, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NON NIL DEP",      None,   1000, 1499, None),
    (19, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "SATP",                  None,   1000, 1499, None),
    (20, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NIL DEP",          None,   1500, None, None),
    (21, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "COMP-NON NIL DEP",      None,   1500, None, None),
    (22, "PCV",  "PCV TAXI (<=6+1, NCB)",    "", "SATP",                  None,   1500, None, None),
    (23, "PCV",  "PCV TAXI STATE CAPITAL (6+1)", "INNOVA/CRYSTA/HYCROSS/SCORPIO/BOLERO", "COMP-NIL DEP",     None, None, None, None),
    (24, "PCV",  "PCV TAXI STATE CAPITAL (6+1)", "INNOVA/CRYSTA/HYCROSS/SCORPIO/BOLERO", "COMP-NON NIL DEP", None, None, None, None),
    (25, "PCV",  "PCV TAXI STATE CAPITAL (6+1)", "INNOVA/CRYSTA/HYCROSS/SCORPIO/BOLERO", "SATP",             None, None, None, None),
    (26, "PCV",  "SCHOOL BUS (18 seater+)",  "", "COMP",                  None,   None, None, None),
    (27, "PCV",  "SCHOOL BUS (18 seater+)",  "", "SATP",                  None,   None, None, None),
]

# PV SATP: fuel groups -> (col, cc_min, cc_max). Bands are INCLUSIVE on both
# ends, so the "above 1500" column starts at 1501 (else cc=1500 is ambiguous).
SATP_SPECS = [
    (4, "PETROL/HYBRID/EV", 0,    1000),
    (5, "PETROL/HYBRID/EV", 1001, 1500),
    (6, "PETROL/HYBRID/EV", 1501, None),
    (7, "DIESEL/CNG/LPG",   0,    1000),
    (8, "DIESEL/CNG/LPG",   1001, 1500),
    (9, "DIESEL/CNG/LPG",   1501, None),
]


def _age_kw(age):
    if not age:
        return {}
    lo, hi = age
    return {"age_min": lo, "age_max": ("" if hi is None else hi)}


def _ccs(cmin, cmax):
    return {"cc_min": ("" if cmin is None else cmin), "cc_max": ("" if cmax is None else cmax)}


def _geo_cluster(label):
    # cluster labels ("MH - M") don't carry a state name; fold via the insurer's
    # RTO cluster map (single-state clusters only — multi-state stays blank and
    # the resolver matches states through CLUSTER2STATES instead)
    return {"geo_kind": "RTO_CLUSTER", "geo_label": label,
            "canonical_state": to_state(label) or cluster_state(label)}


def _geo_state(label):
    return {"geo_kind": "STATE_OR_CITY", "geo_label": label,
            "canonical_state": to_state(label) or (str(label).strip().upper() if label else "")}


def extract_rows():
    wb = openpyxl.load_workbook(FILE, data_only=True)
    rate_rows, elig_rows, warnings = [], [], []
    seq = {"GCV": 0, "PVTCAR": 0, "PCV": 0, "MISD": 0, "HE": 0, "DECL": 0}

    def nid(k):
        seq[k] += 1
        return f"SBI-{k}-{seq[k]:04d}"

    # ---------------------------------------------- GCV sheet (GCV + Pvt Car COMP/SAOD)
    ws = wb[SRC_GCV]
    for r in range(6, 83):
        cluster = _cell(ws, r, 3)
        state = _cell(ws, r, 4)
        if not (cluster and str(cluster).strip()):
            continue
        cluster = str(cluster).strip()
        for col, cat, sub, mk, pol, age, cmin, cmax in GCV_SPECS:
            val = num(_cell(ws, r, col))
            if val is None:
                continue
            k = "PVTCAR" if cat == "PVT_CAR" else "GCV"
            geo = _geo_cluster(cluster)
            # canonical_state: prefer the sheet's own State column when present
            if state and str(state).strip():
                geo["canonical_state"] = to_state(state) or str(state).strip().upper()
            rate_rows.append(row(
                catalog_id=nid(k), source_rule_id=f"{SRC_GCV}!{get_column_letter(col)}{r}",
                rule_type="RATE", effect="RATE", pay_in_pct=val, applies_on="NET",
                category=cat, sub_segment=sub, make=mk, policy_type=pol,
                **_ccs(cmin, cmax), **_age_kw(age), **geo,
                source_sheet=SRC_GCV, source_cell=f"{get_column_letter(col)}{r}",
                source_text=sub + (f" / {mk}" if mk else ""),
                confidence=1.0, review_status="PENDING",
            ))

    # ---------------------------------------------- PCV sheet (MISD + PCV)
    ws = wb[SRC_PCV]
    for r in range(5, 76):
        state = _cell(ws, r, 2)
        if not (state and str(state).strip()):
            continue
        state = str(state).strip()
        for col, cat, sub, mk, pol, age, cmin, cmax, fuel in PCV_SPECS:
            val = num(_cell(ws, r, col))
            if val is None:
                continue
            k = "MISD" if cat == "MISD" else "PCV"
            subseg = sub if not fuel else f"{sub} [{fuel}]"
            rate_rows.append(row(
                catalog_id=nid(k), source_rule_id=f"{SRC_PCV}!{get_column_letter(col)}{r}",
                rule_type="RATE", effect="RATE", pay_in_pct=val, applies_on="NET",
                category=cat, sub_segment=subseg, make=mk, policy_type=pol,
                **_ccs(cmin, cmax), **_age_kw(age), **_geo_state(state),
                source_sheet=SRC_PCV, source_cell=f"{get_column_letter(col)}{r}",
                source_text=subseg + (f" / {mk}" if mk else ""),
                confidence=1.0, review_status="PENDING",
            ))
        # Pvt Car COMP (col AB=28, PO on OD) + SAOD (col AC=29), state-keyed.
        # (this state-level Pvt Car table sits on the PCV sheet and was missed.)
        for col, pol, applies in ((28, "COMP", "OD"), (29, "SAOD", "OD")):
            val = num(_cell(ws, r, col))
            if val is None:
                continue
            rate_rows.append(row(
                catalog_id=nid("PVTCAR"), source_rule_id=f"{SRC_PCV}!{get_column_letter(col)}{r}",
                rule_type="RATE", effect="RATE", pay_in_pct=val, applies_on=applies,
                category="PVT_CAR", sub_segment="PVT CAR", make="", policy_type=pol,
                **_geo_state(state),
                source_sheet=SRC_PCV, source_cell=f"{get_column_letter(col)}{r}",
                source_text=f"Pvt Car {pol} (PO on OD) @ {state}",
                confidence=1.0, review_status="PENDING",
            ))

    # ---------------------------------------------- PV SATP (Pvt Car SATP)
    ws = wb[SRC_SATP]
    n_zero = 0
    for r in range(5, 68):
        state = _cell(ws, r, 2)
        cluster = _cell(ws, r, 3)
        if not (cluster and str(cluster).strip()):
            continue
        cluster = str(cluster).strip()
        for col, fuel, cmin, cmax in SATP_SPECS:
            val = num(_cell(ws, r, col))
            if val is None:
                continue
            if val == 0:
                n_zero += 1
            geo = _geo_cluster(cluster)
            if state and str(state).strip():
                geo["canonical_state"] = to_state(state) or str(state).strip().upper()
            rate_rows.append(row(
                catalog_id=nid("PVTCAR"), source_rule_id=f"{SRC_SATP}!{get_column_letter(col)}{r}",
                rule_type="RATE", effect="RATE", pay_in_pct=val, applies_on="NET",
                category="PVT_CAR", sub_segment=f"PVT CAR SATP [{fuel}]", make="",
                policy_type="SATP", **_ccs(cmin, cmax), **geo,
                source_sheet=SRC_SATP, source_cell=f"{get_column_letter(col)}{r}",
                source_text=f"Pvt Car SATP {fuel} cc {cmin}-{cmax if cmax else '+'}",
                confidence=1.0,
                # a 0% SATP cell is ambiguous (zero payout vs not-offered) -> flag
                review_status=("NEEDS_REVIEW" if val == 0 else "PENDING"),
            ))

    if n_zero:
        warnings.append({
            "scope": "sheet", "cell": SRC_SATP,
            "issue": f"{n_zero} PV SATP cells carry a 0% pay-in (kept as RATE 0, flagged "
                     "review_status=NEEDS_REVIEW); confirm whether 0 means zero payout or "
                     "'not offered' (decline).",
        })
    # SBI-3: header note 'PVT SATP: Grid would be lesser by 3% for vehicle age 1-9 Years'
    # is a value modifier not applied to the stored base rates.
    warnings.append({
        "scope": SRC_SATP, "cell": "header note",
        "issue": "PV SATP base rates are for the base case; per the sheet note the grid is "
                 "3% LOWER for vehicle age 1-9 years — this age adjustment is NOT applied to "
                 "the stored pay_in_pct. Apply the -3% delta downstream for age 1-9.",
    })

    # ---------------------------------------------- PV SAOD Highend (make/model cluster)
    ws = wb[SRC_HE]
    for r in range(5, 590):
        cluster = _cell(ws, r, 1)
        make = _cell(ws, r, 3)
        model = _cell(ws, r, 4)
        if not (make and str(make).strip()):
            continue
        cluster = str(cluster).strip() if cluster else ""
        make = str(make).strip()
        model = str(model).strip() if model else ""
        for col, pol in [(5, "COMP (PACKAGE)"), (6, "SAOD")]:
            val = num(_cell(ws, r, col))
            if val is None:
                continue
            rate_rows.append(row(
                catalog_id=nid("HE"), source_rule_id=f"{SRC_HE}!{get_column_letter(col)}{r}",
                rule_type="RATE", effect="RATE", pay_in_pct=val, applies_on="NET",
                category="PVT_CAR", sub_segment="PVT CAR HIGHEND", make=make, model=model,
                policy_type=pol, **_geo_cluster(cluster),
                source_sheet=SRC_HE, source_cell=f"{get_column_letter(col)}{r}",
                source_text=f"Highend {make} {model} ({pol})",
                confidence=1.0, review_status="PENDING",
            ))

    # ---------------------------------------------- Pvt Car declined make/models (ELIGIBILITY)
    ws = wb[SRC_DECL]
    for r in range(3, 100):
        makes = _cell(ws, r, 2)
        models = _cell(ws, r, 3)
        if not (makes and str(makes).strip()):
            continue
        models = str(models).strip() if models else "All Models & Variants"
        for mk in str(makes).split(","):
            mk = mk.strip()
            if not mk:
                continue
            elig_rows.append(row(
                catalog_id=nid("DECL"), source_rule_id=f"{SRC_DECL}!B{r}",
                rule_type="ELIGIBILITY", effect="DECLINE",
                category="PVT_CAR", sub_segment="PVT CAR", make=mk, model=models,
                geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
                reason=f"Declined make/model (Pvt Car): {mk} — {models}",
                source_sheet=SRC_DECL, source_cell=f"B{r}",
                source_text=f"{makes} | {models}",
                confidence=1.0, review_status="PENDING",
            ))

    # ---------------------------------------------- GCV note declines (condition block)
    def gdecl(sub, make, reason, cell, text, pol=""):
        elig_rows.append(row(
            catalog_id=nid("DECL"), source_rule_id=f"{SRC_GCV}!{cell}",
            rule_type="ELIGIBILITY", effect="DECLINE", category="GCV",
            sub_segment=sub, make=make, policy_type=pol,
            geo_kind="PAN_INDIA", geo_label="PAN_INDIA",
            reason=reason, source_sheet=SRC_GCV, source_cell=cell,
            source_text=text, confidence=1.0, review_status="PENDING"))

    # Pan-India full-segment declines (D75/76/77)
    for sub, cell in (("GCV 3.5-5.0T", "D75"), ("GCV 5.0-7.5T", "D76"), ("GCV 7.5-12T", "D77")):
        gdecl(sub, "", f"{sub}: Pan India decline (all clusters, all make/model)", cell,
              f'* For "{sub}" :- Pan India decline all cluster all make model.')
    # GCV all-tonnage declined makes (D67) — one row per named make, any GCV sub_segment
    for mk in ("Volvo", "MAN", "AMW", "Mercedes", "Eicher", "Scania", "Isuzu",
               "Bharat Benz", "Hyundai"):
        gdecl("", mk, f"Declined make for all GCV tonnage: {mk}", "D67",
              "GCV all tonnage declined makes: imported makes & Volvo, MAN, AMW, Mercedes, "
              "Eicher, Scania, Isuzu, Bharat Benz, Hyundai; Electric fuel declined for all GCV")
    # Cluster-specific GCV declines (D68/71/73/78/79/80) are detailed RTO-cluster lists —
    # surfaced as warnings for human review rather than guessed into rules.
    warnings.append({"scope": SRC_GCV, "cell": "D68:D80",
                     "issue": "GCV per-segment declined-RTO-cluster lists (GCV 3W, Upto-2.5T, "
                              "2.5-3.5T, 12-20T, 20-40T, >40T) are not yet parsed into rules — "
                              "review the condition block and confirm the cluster declines"})

    return rate_rows, elig_rows, warnings


if __name__ == "__main__":
    rate_rows, elig_rows, warnings = extract_rows()
    from collections import Counter
    by_cat = Counter(r["category"] for r in rate_rows + elig_rows)
    print(f"4W+CV extraction: {len(rate_rows)} RATE, {len(elig_rows)} ELIGIBILITY")
    for cat, n in sorted(by_cat.items()):
        print(f"  {cat:10} {n}")
    for s in rate_rows[:3]:
        print("  e.g.", s["catalog_id"], s["category"], s["sub_segment"], s["make"],
              s["policy_type"], "=", s["pay_in_pct"], "@", s["geo_label"], s["source_cell"])
